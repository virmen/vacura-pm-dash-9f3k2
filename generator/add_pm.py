#!/usr/bin/env python3
"""Helper-Skript: legt einen neuen Praxismanager im Excel an.

Erzeugt Token, schreibt in PM-Stammdaten-Sheet UND Daten-Tab.

Aufruf:
    cd ~/Code/Claude/Github/pm-dashboards/v2
    python3 add_pm.py

Danach: Excel wird automatisch gespeichert. Den Rest übernimmt generate.py beim
nächsten Lauf (PM-Liste wird dynamisch aus PM-Stammdaten gelesen).
"""
import os
import secrets
import sys
import openpyxl

EXCEL = os.environ.get('EXCEL_PATH') or os.path.expanduser(
    '~/Code/Claude/Github/pm-dashboards/PM_Gehaltsmodell_v18.xlsx')


def ask(prompt, default=None, type_cast=str, validator=None):
    """Fragt User, mit Default + Validierung."""
    while True:
        suffix = f' [{default}]' if default is not None else ''
        raw = input(f'  {prompt}{suffix}: ').strip()
        if not raw and default is not None:
            return default
        if not raw:
            print('    ⚠️  Pflichtfeld, bitte ausfüllen.')
            continue
        try:
            val = type_cast(raw)
        except (ValueError, TypeError):
            print(f'    ⚠️  Ungültig — erwartet wird {type_cast.__name__}.')
            continue
        if validator and not validator(val):
            continue
        return val


def main():
    print('=' * 64)
    print('  Neuen PM anlegen — Excel-Wizard')
    print('=' * 64)
    print(f'Excel: {EXCEL}\n')

    if not os.path.exists(EXCEL):
        print(f'❌ Excel nicht gefunden: {EXCEL}')
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL)

    if 'PM-Stammdaten' not in wb.sheetnames:
        print('❌ Sheet "PM-Stammdaten" fehlt. Lege es zuerst an (siehe METHODE.md Abschnitt 7).')
        sys.exit(1)
    ws_pm = wb['PM-Stammdaten']
    ws_d = wb['Daten']

    existing = []
    for r in range(6, 50):
        n = ws_pm.cell(row=r, column=1).value
        if isinstance(n, str) and isinstance(ws_pm.cell(row=r, column=2).value, (int, float)):
            existing.append((r, n.strip()))
    existing_names = {n for _, n in existing}
    print(f'Bisherige PMs: {", ".join(n for _, n in existing) or "(keine)"}\n')

    def name_validator(v):
        if v in existing_names:
            print(f'    ⚠️  "{v}" existiert bereits. Anderen Namen wählen.')
            return False
        return True

    name = ask('Name (z.B. "Sophie")', validator=name_validator)
    wochenstd = ask('Wochenstunden', default=40, type_cast=int)
    pm_std_bundle = ask('PM-Std-Bundle (Summe aller PM-Stunden im selben Bundle)', type_cast=int)
    mindestgehalt = ask('Mindestgehalt €/Jahr (üblich 40000-45000)', default=45000, type_cast=int)
    startdatum = ask('Startdatum (YYYY-MM-DD, leer wenn vor 6+ Monaten gestartet)', default='')
    bundle_standorte = ask('Bundle-Standorte (komma-Liste, z.B. "Spandau, Mitte")')
    bundle_pms_raw = ask('Bundle-PMs (komma-Liste der PM-Namen im selben Bundle, inkl. neuer PM)')
    farbe = ask('Farbe (Hex, z.B. "#9C27B0")', default='#0D595A')
    stufe_vorquartal = ask('Stufe Vorquartal (meist 1 für neuen PM)', default=1, type_cast=int,
                           validator=lambda v: 1 <= v <= 6)

    token = secrets.token_hex(16)

    print('\n' + '=' * 64)
    print('Zusammenfassung:')
    print(f'  Name:               {name}')
    print(f'  Wochenstunden:      {wochenstd} h')
    print(f'  PM-Std Bundle:      {pm_std_bundle} h')
    print(f'  Mindestgehalt:      {mindestgehalt} EUR / Jahr')
    print(f'  Startdatum:         {startdatum or "(keines)"}')
    print(f'  Bundle-Standorte:   {bundle_standorte}')
    print(f'  Bundle-PMs:         {bundle_pms_raw}')
    print(f'  Farbe:              {farbe}')
    print(f'  Token (URL):        {token}')
    print(f'  Stufe Vorquartal:   {stufe_vorquartal}')
    print('=' * 64)
    confirm = input('Eintragen? [j/N]: ').strip().lower()
    if confirm not in ('j', 'ja', 'y', 'yes'):
        print('❌ Abgebrochen, keine Änderung.')
        sys.exit(0)

    next_row_pm = max((r for r, _ in existing), default=5) + 1
    while ws_pm.cell(row=next_row_pm, column=1).value:
        next_row_pm += 1
    ws_pm.cell(row=next_row_pm, column=1, value=name)
    ws_pm.cell(row=next_row_pm, column=2, value=wochenstd)
    ws_pm.cell(row=next_row_pm, column=3, value=pm_std_bundle)
    ws_pm.cell(row=next_row_pm, column=4, value=mindestgehalt)
    ws_pm.cell(row=next_row_pm, column=5, value=startdatum or None)
    ws_pm.cell(row=next_row_pm, column=6, value=bundle_standorte)
    ws_pm.cell(row=next_row_pm, column=7, value=bundle_pms_raw)
    ws_pm.cell(row=next_row_pm, column=8, value=farbe)
    ws_pm.cell(row=next_row_pm, column=9, value=token)
    ws_pm.cell(row=next_row_pm, column=10, value=True)

    next_row_d = 5
    while ws_d.cell(row=next_row_d, column=2).value:
        next_row_d += 1
    ws_d.cell(row=next_row_d, column=1, value=startdatum or None)
    ws_d.cell(row=next_row_d, column=2, value=name)
    ws_d.cell(row=next_row_d, column=3, value=wochenstd)
    ws_d.cell(row=next_row_d, column=4, value=pm_std_bundle)
    ws_d.cell(row=next_row_d, column=5, value=stufe_vorquartal)
    ws_d.cell(row=next_row_d, column=6, value=mindestgehalt)

    wb.save(EXCEL)

    print(f'\n✅ {name} eingetragen.')
    print(f'   PM-Stammdaten Zeile {next_row_pm}')
    print(f'   Daten-Tab Zeile {next_row_d}')
    print(f'   Token: {token}')
    print('\n🔗 Dashboard-URL (nach nächstem Deploy):')
    print(f'   https://virmen.github.io/vacura-pm-dash-9f3k2/{name.lower()}-{token}.html')
    print('\n📋 Nächste Schritte:')
    print('   1. (optional) Q-Werte für Bewertungs-Quartal manuell in Daten-Tab Spalten 7-10 nachtragen')
    print('   2. (optional) Zufriedenheits-Score in Spalten 11-13 nachtragen (sobald Q-Umfrage da)')
    print('   3. python3 generate.py — Dashboard lokal generieren zur Vorschau')
    print('   4. Sync ins Deploy-Repo + push (oder warte auf nächste Daily-Action 06:00 Berlin)')


if __name__ == '__main__':
    main()
