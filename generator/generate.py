#!/usr/bin/env python3
"""
Generator für PM-Dashboard-HTMLs v2 (für PM_Gehaltsmodell.xlsx).

Liest v18 Excel + rechnet KPIs selbst aus (keine Excel-Cache-Voraussetzung).
Generiert eine HTML pro PM mit dem 9-Block Roten Faden.

Usage: python3 generate.py
"""
import openpyxl, os, sys, html, json
from datetime import datetime, date

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.environ.get('EXCEL_PATH') or os.path.expanduser('~/Code/Claude/Github/pm-dashboards/PM_Gehaltsmodell.xlsx')
OUT_DIR = os.environ.get('OUT_DIR') or os.path.expanduser('~/Code/Claude/Github/pm-dashboards/code')

# === Abrechnungs-Systematik (Valentin, 22.07.2026 — gegen MediFox Q1+Q2 backtestet) ===
# Einheitspreis pro Zeitintervall (15 min): jede Behandlung wird mit Kalenderdauer/15
# PLUS 1 ZI Vor-/Nachbereitung abgerechnet. Reproduziert die GKV-Staffel exakt:
# 30 min = 3 ZI = 56,93 · 45 min = 4 ZI = 75,91 · 60 min = 5 ZI = 94,89.
# Ausnahmen: thermische Anwendung/KT/WT immer 8,51 €; Hausbesuchspauschale +27,56 (nach Faktor-Logik: NICHT ×Faktor).
# PKV (verordnungstyp 2) ×2,0 · Selbstzahler (3) ×1,7 — empirisch kalibriert (Monatsumsatz-Report V3).
# Wirkt auf IST (termin_umsatz) und den Hebel-Faktor (1 %-Pkt PKV ≈ (PKV_FAKTOR-1)*100 % Umsatz).
ZI_PREIS = 18.98
THERMISCH_PREIS = 8.51
PKV_FAKTOR = 2.0
SZ_FAKTOR = 1.7
HB_PAUSCHALE = 27.56
# GKV-Schiedsspruch: +4,11 % auf alle Sätze für Behandlungen ab 01.07.2026
ERHOEHUNG_AB = '2026-07-01'
ERHOEHUNG_FAKTOR = 1.0411
# Je VO: 10 € Verordnungsblattgebühr; je Blanko-VO zusätzlich 98,59 € Versorgungspauschale
# (Pos. 54503) — zugeordnet dem Quartal des letzten VO-Termins.
VO_BLATTGEBUEHR = 10.00
BLANKO_PAUSCHALE = 98.59

def satz_faktor(iso_datum):
    """+4,11 % ab 01.07.2026 (Behandlungsdatum)."""
    return ERHOEHUNG_FAKTOR if str(iso_datum) >= ERHOEHUNG_AB else 1.0

# Stufen
# STUFEN-Default: gilt für Tests + als Fallback.
# Excel-Tab `Stufentabelle` ist die Quelle der Wahrheit, wird beim Excel-Load eingelesen
# und überschreibt diese Defaults. Bei jährlichem Schwellen-Review nur Excel ändern.
STUFEN = [
    {'n': 1, 'name': 'Basis',         'zufr': 5.0, 'zulage': 0.00, 'eur60': 61.17},
    {'n': 2, 'name': 'Gut',           'zufr': 6.0, 'zulage': 0.11, 'eur60': 66.54},
    {'n': 3, 'name': 'Stark',         'zufr': 7.0, 'zulage': 0.22, 'eur60': 72.64},
    {'n': 4, 'name': 'Sehr stark',    'zufr': 8.0, 'zulage': 0.33, 'eur60': 77.28},
    {'n': 5, 'name': 'Exzellent',     'zufr': 8.5, 'zulage': 0.44, 'eur60': 84.11},
    {'n': 6, 'name': 'Herausragend',  'zufr': 8.5, 'zulage': 0.55, 'eur60': 89.48},
]

# Maximaler Stufen-Sprung pro Quartal (Übersprungs-Limit gegenüber Vorquartal-Stufe).
# Gilt für tats_stufe; rechn_stufe (rein aus eur60+zufr) bleibt davon unberührt
# und wird im Live-Block zur Motivation gezeigt.
MAX_STUFEN_SPRUNG = 1

def load_stufen_aus_excel(workbook):
    """Liest Stufen-Schwellen aus Excel-Tab 'Stufentabelle' und überschreibt das globale STUFEN.

    Excel-Schema: Spalte 1=Stufe-Nr, 4=Zufr-Schwelle, 5=Zulage % (als Dezimalzahl), 6=€/60min-Schwelle.
    Stufen-Namen bleiben aus den Defaults (sind reine UI-Labels, nicht im Excel).
    """
    global STUFEN
    try:
        ws = workbook['Stufentabelle']
    except KeyError:
        return  # Tab nicht da → Defaults behalten
    name_map = {s['n']: s['name'] for s in STUFEN}
    neu = []
    for r in range(2, 12):
        n = ws.cell(row=r, column=1).value
        if not isinstance(n, (int, float)) or n < 1 or n > 6:
            continue
        zufr = ws.cell(row=r, column=4).value
        zulage = ws.cell(row=r, column=5).value
        eur60 = ws.cell(row=r, column=6).value
        if zufr is None or zulage is None or eur60 is None:
            continue
        neu.append({'n': int(n), 'name': name_map.get(int(n), f'Stufe {int(n)}'),
                    'zufr': float(zufr), 'zulage': float(zulage), 'eur60': float(eur60)})
    if len(neu) == 6:
        STUFEN = sorted(neu, key=lambda s: s['n'])

# KPI-Level für Wege-Block
KPI_LEVELS = {
    'Auslastung': {
        'low':   {'text': 'niedrig',   'range': 'bis 84 %'},
        'mid':   {'text': 'mittel',    'range': '85–92 %'},
        'high':  {'text': 'hoch',      'range': '93–95 %'},
        'vhigh': {'text': 'sehr hoch', 'range': '96 %+'},
    },
    'PKV-Quote': {
        'low':   {'text': 'gering',    'range': 'bis 10 %'},
        'mid':   {'text': 'mittel',    'range': '11–20 %'},
        'high':  {'text': 'hoch',      'range': '21–30 %'},
        'vhigh': {'text': 'sehr hoch', 'range': 'über 30 %'},
    },
    'Krankheit': {
        'low':   {'text': 'wenig',     'range': 'bis 15 Tg./a'},
        'mid':   {'text': 'normal',    'range': '16–20 Tg./a'},
        'high':  {'text': 'höher',     'range': '21–25 Tg./a'},
        'vhigh': {'text': 'sehr viel', 'range': 'über 25 Tg./a'},
    },
}

STUFEN_WEGE = {
    1: {'type': 'text', 'text': 'Standort im Aufbau oder nach Veränderungen — Grundlagen werden gelegt.'},
    2: {'type': 'text', 'text': 'Solide Auslastung, Team läuft stabil.'},
    3: {'type': 'wege', 'wege': [
        [('Auslastung','vhigh'), ('PKV-Quote','low'), ('Krankheit','high')],
        [('Auslastung','high'),  ('PKV-Quote','low'), ('Krankheit','high')],
        [('Auslastung','mid'),   ('PKV-Quote','low'), ('Krankheit','low')],
    ]},
    4: {'type': 'wege', 'wege': [
        [('Auslastung','vhigh'), ('PKV-Quote','mid'),  ('Krankheit','high')],
        [('Auslastung','high'),  ('PKV-Quote','low'),  ('Krankheit','low')],
        [('Auslastung','mid'),   ('PKV-Quote','high'), ('Krankheit','low')],
    ]},
    5: {'type': 'wege', 'wege': [
        [('Auslastung','vhigh'), ('PKV-Quote','high'),  ('Krankheit','high')],
        [('Auslastung','mid'),   ('PKV-Quote','vhigh'), ('Krankheit','low')],
    ]},
    6: {'type': 'text', 'text': 'Konstant Top-Ergebnisse über mehrere Quartale auf Stufe-5-Niveau.'},
}

def render_wege_block(next_stufe_num):
    """Rendert Kombinations-Wege für die Zielstufe."""
    if next_stufe_num not in STUFEN_WEGE:
        return ''
    cfg = STUFEN_WEGE[next_stufe_num]
    if cfg['type'] == 'text':
        return f'<div class="weg-text">{cfg["text"]}</div>'
    weg_labels = ['Weg A','Weg B','Weg C','Weg D']
    wege_html = ''
    for i, kombi in enumerate(cfg['wege']):
        vars_html = ''
        for kpi_name, level in kombi:
            level_info = KPI_LEVELS[kpi_name][level]
            vars_html += (
                '<div class="weg-var">'
                f'<div class="weg-var-label">{kpi_name}</div>'
                '<div class="weg-chip-track">'
                f'<div class="weg-chip weg-chip-{level}">'
                f'<span class="weg-chip-text">{level_info["text"]}</span>'
                f'<span class="weg-chip-range">{level_info["range"]}</span>'
                '</div>'
                '</div>'
                '</div>'
            )
        wege_html += (
            '<div class="weg-card">'
            f'<div class="weg-head"><span class="weg-head-label">{weg_labels[i]}</span><span class="weg-head-tag">Kombination</span></div>'
            f'<div class="weg-vars">{vars_html}</div>'
            '</div>'
        )
    return f'<div class="wege-grid">{wege_html}</div>'


# PMS wird zur Laufzeit aus dem Excel-Tab 'PM-Stammdaten' geladen
# (siehe load_pms_from_excel). Default = leer; bei direkter Nutzung der
# Module (z.B. Tests) ohne Excel bleibt PMS leer.
PMS = []
TOKENS = {}   # Name → 32-hex-Token, aus PM-Stammdaten gelesen

def load_pms_from_excel(workbook):
    """Liest PM-Stammdaten aus Excel-Tab 'PM-Stammdaten' und befüllt globales PMS + TOKENS.

    Auto-Fix-Verhalten (Excel-first-Workflow):
    - **Token leer** → automatisch generieren (`secrets.token_hex(16)`) und ins Excel zurückschreiben
    - **Farbe leer** → Default `#0D595A` setzen
    - **Aktiv leer** → True (aktiv) annehmen
    - **PM in Daten-Tab fehlt** → Stammdaten automatisch dorthin spiegeln (mit Stufe Vorquartal=1)

    Returns: True wenn Workbook geändert wurde (Caller muss wb.save() aufrufen).

    Schema (Sheet 'PM-Stammdaten'):
    Spalte 1=Name, 2=Wochenstd, 3=PM-Std Bundle, 4=Mindestgehalt, 5=Startdatum,
    6=Bundle-Standorte, 7=Bundle-PMs (komma-Liste), 8=Farbe (#RRGGBB),
    9=URL-Token (32-hex), 10=Aktiv (bool).

    Inaktive PMs (Aktiv=False) werden übersprungen — Stammdaten bleiben aber zum Lesen erhalten.
    """
    import secrets
    global PMS, TOKENS
    if 'PM-Stammdaten' not in workbook.sheetnames:
        return False
    ws = workbook['PM-Stammdaten']
    changed = False

    pms = []
    tokens = {}
    for r in range(6, 50):
        name = ws.cell(row=r, column=1).value
        wochenstd = ws.cell(row=r, column=2).value
        # Nur Zeilen mit echtem PM-Namen + numerischer Wochenstunden-Zelle akzeptieren
        if not name or not isinstance(name, str) or not isinstance(wochenstd, (int, float)):
            continue
        name = name.strip()

        # Aktiv: leer → True
        aktiv = ws.cell(row=r, column=10).value
        if aktiv is None:
            ws.cell(row=r, column=10, value=True)
            aktiv = True
            changed = True
        if aktiv is False or (isinstance(aktiv, str) and aktiv.lower() in ('nein', 'no', 'false', '0')):
            continue

        # Farbe: leer → Default
        farbe = ws.cell(row=r, column=8).value
        if not farbe:
            farbe = '#0D595A'
            ws.cell(row=r, column=8, value=farbe)
            changed = True

        # Token: leer → automatisch erzeugen + zurückschreiben
        token = ws.cell(row=r, column=9).value
        if not token:
            token = secrets.token_hex(16)
            ws.cell(row=r, column=9, value=token)
            changed = True
            print(f'  🔑 Token für {name} automatisch erzeugt: {token}')

        bundle_pms_raw = ws.cell(row=r, column=7).value or ''
        # Stammdaten komplett aus PM-Stammdaten-Sheet — kein Daten-Tab-Sync mehr nötig
        pm_std_bundle = ws.cell(row=r, column=3).value
        mindestgehalt = ws.cell(row=r, column=4).value
        startdatum = ws.cell(row=r, column=5).value

        pms.append({
            'name': name,
            'color': farbe,
            'wochenstd': wochenstd,
            'pm_std_bundle': pm_std_bundle,
            'mindestgehalt': mindestgehalt,
            'startdatum': startdatum,
            'bundle_pms': [p.strip() for p in str(bundle_pms_raw).split(',') if p.strip()],
            'bundle_standorte': ws.cell(row=r, column=6).value or '',
        })
        tokens[name] = str(token).strip()

    PMS = pms
    TOKENS = tokens
    return changed

def th_kumuliert(n_th):
    """TH-Zulage kumuliert für n TH-Äquivalente."""
    cum = 0
    for i in range(1, n_th + 1):
        if i <= 4: cum += 250
        elif i <= 9: cum += 400
        else: cum += 700
    return cum

_BUNDLE_VZAE_CACHE = {}

def bundle_brutto_vzae(bundle_standorte, stichtag):
    """Bundle-Größe in 30h-VZÄ aus Brutto-Vertragsstunden am Stichtag (v1-Methode).

    Summiert die am Stichtag gültigen Vertrags-Wochenstunden (arbeitszeit_gruppen)
    aller am Stichtag beschäftigten Therapeut:innen der Bundle-Filialen, ÷ 30.
    Entspricht der Vertragsdefinition in METHODE.md 5.2 (Brutto-Wochenstunden,
    kein LZ-Abzug). Return: int VZÄ, oder None wenn NocoDB nicht erreichbar
    (Caller fällt dann auf den vstd_ber-Proxy zurück).
    """
    iso = stichtag.isoformat()
    standorte = tuple(sorted(s.strip().lower().replace(' ', '_')
                             for s in bundle_standorte.split(',') if s.strip()))
    key = (standorte, iso)
    if key in _BUNDLE_VZAE_CACHE:
        return _BUNDLE_VZAE_CACHE[key]
    try:
        ma = _fetch_all('mc934lbrlg7w6e1')
    except Exception as e:
        print(f"    ⚠️  NocoDB für Bundle-VZÄ nicht erreichbar ({str(e)[:120]}) — Fallback vstd_ber-Proxy")
        return None
    weekly = 0.0
    for m in ma:
        if not m.get('is_therapeut'):
            continue
        if 'Online' in f"{m.get('vorname','')} {m.get('nachname','')}":
            continue
        if not any(f in standorte for f in (m.get('filialen') or [])):
            continue
        bz = m.get('beschaeftigungszeiten') or []
        aktiv = any((e.get('Von') is None or e['Von'] <= iso) and
                    (e.get('Bis') is None or e['Bis'] >= iso) for e in bz)
        if not aktiv:
            continue
        gruppen = [g for g in (m.get('arbeitszeit_gruppen') or [])
                   if (g.get('GueltigAb') or '0000') <= iso
                   and (not g.get('GueltigBis') or g['GueltigBis'] >= iso)]
        if not gruppen:
            continue
        gruppen.sort(key=lambda g: g.get('GueltigAb') or '', reverse=True)
        weekly += float(gruppen[0].get('StundenProWoche', 0) or 0)
    vzae = round(weekly / 30)
    _BUNDLE_VZAE_CACHE[key] = vzae
    return vzae

def get_cell_val(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if isinstance(v, str) and v.startswith('='):
        # Formel — versuche zu resolven
        return None
    return v

def resolve_formula(ws, r, c, max_depth=3):
    """Löse einfache =X5-Referenzen auf."""
    v = ws.cell(row=r, column=c).value
    if not isinstance(v, str) or not v.startswith('='):
        return v
    # Einfache Referenz wie =G6
    if max_depth == 0: return None
    import re
    m = re.match(r'^=([A-Z]+)(\d+)$', v.strip())
    if m:
        from openpyxl.utils import column_index_from_string
        col = column_index_from_string(m.group(1))
        row = int(m.group(2))
        return resolve_formula(ws, row, col, max_depth-1)
    return None

def _find_qb_row(ws_qb, quartal_label, name):
    """Findet Zeile in Quartals-Bewertungen für (quartal_label, PM-Name).
    Spalten: 1=Quartal (z.B. 'Q1 2026'), 2=PM. Header in Row 7. Daten ab Row 8.
    Returns Row-Nr oder None."""
    for r in range(8, 200):
        q = ws_qb.cell(row=r, column=1).value
        n = ws_qb.cell(row=r, column=2).value
        if isinstance(q, str) and isinstance(n, str) and q.strip() == quartal_label and n.strip() == name:
            return r
    return None


def _q_label_from_date(d):
    """date → 'Q1 2026' / 'Q2 2026' / ..."""
    return f'Q{(d.month - 1) // 3 + 1} {d.year}'


def _previous_q_label_from(quartal_label):
    """'Q2 2026' → 'Q1 2026'; 'Q1 2027' → 'Q4 2026'."""
    parts = quartal_label.split(' ')
    q_num = int(parts[0][1:])
    year = int(parts[1])
    if q_num == 1:
        return f'Q4 {year - 1}'
    return f'Q{q_num - 1} {year}'


def _find_pm_row(ws_daten, name):
    """LEGACY: Findet die Daten-Tab-Row eines PMs per Name.
    Wird nicht mehr genutzt seit Daten-Tab entfernt — bleibt für Backwards-Compat."""
    if ws_daten is None: return None
    for r in range(5, 30):
        v = ws_daten.cell(row=r, column=2).value
        if v and isinstance(v, str) and v.strip() == name:
            return r
    return None

def compute_pm(wb_or_ws, pm, q_label='Q1 2026'):
    """Berechnet PM-Q-Bewertung aus Excel.

    Stammdaten kommen aus pm-Dict (geladen via load_pms_from_excel).
    Q-Werte (IST, Vstd, Abw, Zufriedenheit) kommen aus Quartals-Bewertungen-Sheet.

    Args:
        wb_or_ws: Workbook (neuer Standard) ODER Daten-Tab (Legacy für Backwards-Compat).
        pm: dict mit 'name', 'wochenstd', 'pm_std_bundle', 'mindestgehalt', 'startdatum',
            'bundle_standorte', 'color', etc.
        q_label: Quartal-Label ('Q1 2026' = historisch-Default).

    Returns: dict mit pm-Daten + Q-Bewertung, oder None wenn Q-Werte fehlen.
    """
    # Workbook ermitteln
    if hasattr(wb_or_ws, 'sheetnames'):
        wb = wb_or_ws
    else:
        # Legacy: ws_daten übergeben — finde das Workbook
        wb = wb_or_ws.parent
    if 'Quartals-Bewertungen' not in wb.sheetnames:
        print(f"    ⚠️  Sheet 'Quartals-Bewertungen' fehlt, übersprungen")
        return None
    ws_qb = wb['Quartals-Bewertungen']

    # Stammdaten direkt aus pm-Dict (kommt aus PM-Stammdaten)
    wochenstd = pm.get('wochenstd')
    pm_std_bundle = pm.get('pm_std_bundle')
    mindestgehalt = pm.get('mindestgehalt')
    startdatum = pm.get('startdatum')

    if not all([wochenstd, pm_std_bundle, mindestgehalt]):
        print(f"    ⚠️  Unvollständige Stammdaten für {pm['name']}, übersprungen")
        return None

    # Q-Werte aus Quartals-Bewertungen
    qb_row = _find_qb_row(ws_qb, q_label, pm['name'])
    if not qb_row:
        print(f"    ⚠️  Keine Q-Bewertungs-Zeile für ({q_label}, {pm['name']}), übersprungen")
        return None

    # Spalten in Quartals-Bewertungen:
    # 3=Rücken, 4=Komm, 5=eNPS, 6=MediFox, 7=Zufr-Score (berechnet),
    # 8=Vstd, 9=Abw, 10=Feiertage, 11=IST, 12=verfueg, 13=eur60,
    # 14=Rechn-Stufe, 15=Tats-Stufe, 16=Probezeit, 17=Diff
    ruecken = ws_qb.cell(row=qb_row, column=3).value or 0
    komm    = ws_qb.cell(row=qb_row, column=4).value or 0
    enps    = ws_qb.cell(row=qb_row, column=5).value or 0
    vstd_ber = ws_qb.cell(row=qb_row, column=8).value
    abw_ber  = ws_qb.cell(row=qb_row, column=9).value
    feiertage_ber = ws_qb.cell(row=qb_row, column=10).value or 0
    verfueg_sheet = ws_qb.cell(row=qb_row, column=12).value
    ist      = ws_qb.cell(row=qb_row, column=11).value

    # Probezeit prüfen (für Q-Bewertungs-Ende ermittelt)
    from datetime import date as _ddate
    parts = q_label.split(' ')
    q_num_x = int(parts[0][1:])
    year_x = int(parts[1])
    q_end_month = q_num_x * 3
    q_end_day = 31 if q_end_month in (3, 12) else 30
    q_eval_end = _ddate(year_x, q_end_month, q_end_day)
    probezeit_aktiv = is_probezeit(startdatum, q_eval_end)

    # Probezeit-PMs: kein Q-IST/Vstd nötig — gehen automatisch auf Stufe 1, Mindestgehalt-Pfad
    if probezeit_aktiv:
        # Synthetische Werte, damit Render-Logik nicht crasht
        vstd_ber = vstd_ber or 1
        abw_ber  = abw_ber  or 0
        ist      = ist      or 0
    elif not all([vstd_ber, abw_ber, ist]):
        # Reguläre PMs ohne Q-Werte: nicht bewertbar (z.B. Q2 vor Q-End-Routine)
        return None

    # Start-Stufe aus Vorquartal-tats_stufe (für ±1-Deckel)
    prev_q = _previous_q_label_from(q_label)
    prev_row = _find_qb_row(ws_qb, prev_q, pm['name'])
    start_stufe = 1
    if prev_row:
        v = ws_qb.cell(row=prev_row, column=15).value
        if isinstance(v, (int, float)): start_stufe = int(v)

    # Vstd-Bundle (für Bundle-Zulage) — nicht im Q-Bewertungs-Sheet, sondern hochgerechnet
    # aus PM-Std-Bundle und Wochenstunden-Verhältnis. Vereinfachung: vstd_ber als Proxy.
    vstd_bundle = vstd_ber
    
    # Ableitung: verfueg bevorzugt aus der Routine-Spalte 12; Fallback rechnet die
    # Feiertage-Spalte mit ein (Q1-Altzeilen haben Feiertage bereits in Vstd/Abw
    # verrechnet und Spalte 10 leer — dort ist der Abzug 0).
    if isinstance(verfueg_sheet, (int, float)) and verfueg_sheet > 0:
        verfueg = verfueg_sheet
    else:
        verfueg = vstd_ber - abw_ber - feiertage_ber
    eur60 = ist / verfueg
    zufr = ruecken*0.2 + komm*0.2 + enps*0.6

    # Rechn. Stufe
    rechn = 0
    for s in reversed(STUFEN):
        if eur60 >= s['eur60'] and zufr >= s['zufr']:
            rechn = s['n']; break
    
    # Tats. Stufe ±1 Deckel
    if rechn == 0:
        tats = max(start_stufe - 1, 1)
    elif rechn > start_stufe + 1:
        tats = start_stufe + 1
    elif rechn < start_stufe - 1:
        tats = max(start_stufe - 1, 1)
    else:
        tats = rechn
    
    # TH-Äqui für Bundle-Zulage (Stichtagswert — keine 29-Tage-Sperre, siehe Memory).
    # Brutto-Vertragsstunden HEUTE aus NocoDB (v1-Methode, METHODE.md 5.2): die
    # Bundle-Zulage läuft vertraglich monatlich mit der aktuellen Bundle-Größe —
    # nicht mit dem Q-Schnappschuss. Der frühere vstd_ber-Proxy maß zudem wegen
    # LZ-Abzug ~1 VZÄ zu klein.
    stichtag = _ddate.today()
    th_bundle = bundle_brutto_vzae(pm.get('bundle_standorte', ''), stichtag)
    if th_bundle is None:
        th_bundle = round(vstd_bundle / 13 / 30) if vstd_bundle else 0
    th_pm = round(th_bundle * wochenstd / pm_std_bundle) if pm_std_bundle else 0

    # Probezeit wurde oben schon ermittelt (mit korrektem q_eval_end aus q_label)
    # Gehalt
    if probezeit_aktiv:
        tats = 1
        bundle_zulage = 0
    else:
        bundle_zulage = th_kumuliert(th_pm)
    basis = 40000 + bundle_zulage
    stufe_zulage_pct = STUFEN[tats - 1]['zulage']
    gehalt_formel = round(basis * (1 + stufe_zulage_pct) * (wochenstd / 40))
    min_gehalt_anteilig = round(mindestgehalt * wochenstd / 40)
    jahr = max(gehalt_formel, min_gehalt_anteilig)
    monat = round(jahr / 12)
    
    return {
        'name': pm['name'],
        'color': pm['color'],
        'wochenstd': wochenstd,
        'pm_std_bundle': pm_std_bundle,
        'start_stufe': start_stufe,
        'mindestgehalt': mindestgehalt,
        'startdatum': startdatum,
        'vstd_bundle': vstd_bundle,
        'vstd_ber': vstd_ber,
        'abw_ber': abw_ber,
        'ist': ist,
        'ruecken': ruecken, 'komm': komm, 'enps': enps,
        'verfueg': verfueg,
        'eur60': eur60,
        'zufr': zufr,
        'rechn_stufe': rechn,
        'tats_stufe': tats,
        'tats_stufe_name': STUFEN[tats-1]['name'],
        'tats_stufe_zulage_pct': stufe_zulage_pct,
        'th_bundle': th_bundle,
        'th_pm': th_pm,
        'bundle_zulage': bundle_zulage,
        'basis_gehalt': basis,
        'gehalt_formel': gehalt_formel,
        'jahresgehalt': jahr,
        'monatsgehalt': monat,
        'mindest_anteilig': min_gehalt_anteilig,
        'bundle_standorte': pm['bundle_standorte'],
        'bundle_pms': pm['bundle_pms'],
        'probezeit_aktiv': probezeit_aktiv,
    }

def delta_naechste_stufe(pm_data):
    tats = pm_data['tats_stufe']
    if tats >= 6:
        return None
    next_s = STUFEN[tats]
    delta_eur60 = next_s['eur60'] - pm_data['eur60']
    delta_bundle_umsatz_quartal = (next_s['eur60'] - pm_data['eur60']) * pm_data['verfueg']
    delta_bundle_monat = delta_bundle_umsatz_quartal / 3
    # Gehaltsunterschied: next_gehalt - aktuelles_gehalt
    next_zulage = next_s['zulage']
    curr_zulage = pm_data['tats_stufe_zulage_pct']
    next_jahr = round(pm_data['basis_gehalt'] * (1 + next_zulage) * (pm_data['wochenstd'] / 40))
    next_monat = round(next_jahr / 12)
    delta_monat = next_monat - pm_data['monatsgehalt']
    delta_jahr = next_jahr - pm_data['jahresgehalt']
    return {
        'next_stufe': next_s,
        'delta_eur60': delta_eur60,
        'delta_bundle_monat': delta_bundle_monat,
        'delta_gehalt_monat': delta_monat,
        'delta_gehalt_jahr': delta_jahr,
        'next_jahresgehalt': next_jahr,
        'next_monatsgehalt': next_monat,
        # Progress-Percent: wie weit im Gap
        'progress_pct': max(0, min(100, (pm_data['eur60'] - STUFEN[tats-1]['eur60']) /
                                        (next_s['eur60'] - STUFEN[tats-1]['eur60']) * 100)),
    }

def hebel_optionen(pm_data, gap_data, live_kpis):
    """Konkrete Hebel-Werte um die nächste Stufe zu erreichen.

    Faktoren (Standard-Approximation):
      +1 %-Pkt PKV-Quote     ≈ +0,7 % Umsatz
      +1 Termin/Wo (Bundle)  ≈ +0,3 % Umsatz
      −1 %-Pkt Krankenstand  ≈ +0,5 % Umsatz   (1 %-Pkt ≈ 2,3 Tage/TH/Jahr bei 230 Werktagen)
    """
    if not gap_data:
        return None
    if not pm_data.get('eur60') or pm_data['eur60'] <= 0:
        return None   # Probezeit-PMs ohne Q-Werte → keine Hebel
    delta_pct = gap_data['delta_eur60'] / pm_data['eur60'] * 100
    if delta_pct <= 0:
        return None

    F_PKV, F_TERMIN, F_KRANK = (PKV_FAKTOR - 1), 0.3, 0.5
    d_pkv_pkt   = delta_pct / F_PKV
    d_termin_wo = delta_pct / F_TERMIN
    d_krank_pkt = delta_pct / F_KRANK
    d_krank_tage = d_krank_pkt * 2.3

    pkv_now    = (live_kpis or {}).get('pkv_quote') or 0
    krank_now  = (live_kpis or {}).get('krank_tage_pro_th_jahr') or 0
    auslast_now = (live_kpis or {}).get('auslastung') or 0
    anzahl_th  = (live_kpis or {}).get('auslastung_n_th') or 0
    d_termin_pro_th = (d_termin_wo / anzahl_th) if anzahl_th else None

    # 3-Stufen-Plausibilität: realistic | borderline | impossible
    if d_pkv_pkt <= 15:    pkv_lvl = 'realistic'
    elif d_pkv_pkt <= 25:  pkv_lvl = 'borderline'
    else:                  pkv_lvl = 'impossible'

    # Termine: pro-TH-Wert ist die natürliche Plausibilitäts-Einheit.
    # Auch bei voller Auslastung kann der Hebel ziehen — über PKV-Fokus bei rotierenden Patienten.
    pth = d_termin_pro_th or d_termin_wo
    if pth <= 1:    termin_lvl = 'realistic'
    elif pth <= 2:  termin_lvl = 'borderline'
    else:           termin_lvl = 'impossible'

    if not krank_now or krank_now <= 0:  krank_lvl = 'impossible'
    elif d_krank_tage <= krank_now * 0.5: krank_lvl = 'realistic'
    elif d_krank_tage <= krank_now:      krank_lvl = 'borderline'
    else:                                krank_lvl = 'impossible'

    return {
        'delta_pct': delta_pct,
        'd_pkv_pkt': d_pkv_pkt, 'pkv_now': pkv_now, 'pkv_neu': pkv_now + d_pkv_pkt,
        'd_termin_wo': d_termin_wo, 'd_termin_pro_th': d_termin_pro_th, 'anzahl_th': anzahl_th,
        'd_krank_tage': d_krank_tage, 'krank_now': krank_now, 'krank_neu': max(krank_now - d_krank_tage, 0),
        'pkv_lvl': pkv_lvl, 'termin_lvl': termin_lvl, 'krank_lvl': krank_lvl,
    }

# ==================== HTML TEMPLATE ====================
CSS = """
:root {
  --teal: #0D595A;
  --teal-mid: #1a8a8b;
  --teal-light: #e8f4f4;
  --teal-glow: rgba(13, 89, 90, 0.08);
  --orange: #ED7D31;
  --orange-light: #fef0e5;
  --green: #2e7d32;
  --green-light: #e8f5e9;
  --red: #c62828;
  --red-light: #ffebee;
  --bg: #fafafa;
  --white: #ffffff;
  --ink: #1a1a1a;
  --ink-soft: #4a4a4a;
  --muted: #8e8e8e;
  --line: #e8e8e8;
  --line-strong: #d4d4d4;
  --radius: 20px;
  --radius-sm: 12px;
  --shadow-sm: 0 1px 2px rgba(0,0,0,0.04);
  --shadow: 0 4px 24px rgba(0,0,0,0.06);
  --shadow-lg: 0 12px 40px rgba(0,0,0,0.08);
  --font: -apple-system, BlinkMacSystemFont, 'Inter', 'Segoe UI', Roboto, sans-serif;
  --mono: 'SF Mono', Menlo, Monaco, monospace;
}

* { margin: 0; padding: 0; box-sizing: border-box; }
html { scroll-behavior: smooth; }
body {
  font-family: var(--font);
  background: var(--bg);
  color: var(--ink);
  line-height: 1.55;
  font-feature-settings: 'tnum' 1, 'cv11' 1;
  -webkit-font-smoothing: antialiased;
}

.container { max-width: 760px; margin: 0 auto; padding: 32px 20px 80px; }

/* === TOP BAR === */
.topbar {
  display: flex;
  align-items: center;
  justify-content: space-between;
  margin-bottom: 28px;
  padding-bottom: 16px;
  border-bottom: 1px solid var(--line);
}
.topbar-name {
  font-size: 13px;
  color: var(--muted);
  letter-spacing: 0.04em;
  text-transform: uppercase;
  font-weight: 600;
}
.topbar-quartal {
  font-size: 13px;
  color: var(--ink);
  font-weight: 500;
  padding: 4px 12px;
  background: var(--teal-light);
  border-radius: 999px;
  color: var(--teal);
}

/* === PAGE TITLE === */
.page-title {
  font-size: 32px;
  font-weight: 800;
  letter-spacing: -0.02em;
  margin-bottom: 4px;
  line-height: 1.1;
}
.page-subtitle {
  font-size: 15px;
  color: var(--muted);
  margin-bottom: 40px;
}

/* === BLOCK === */
.block { margin-bottom: 36px; }
.block-label {
  display: flex; align-items: center; gap: 8px;
  font-size: 11px;
  font-weight: 700;
  letter-spacing: 0.1em;
  text-transform: uppercase;
  color: var(--muted);
  margin-bottom: 14px;
}
.block-label::before {
  content: ''; width: 6px; height: 6px; border-radius: 50%;
  background: var(--teal);
}
.block-title {
  font-size: 19px;
  font-weight: 700;
  color: var(--ink);
  margin-bottom: 16px;
  letter-spacing: -0.01em;
}
.block-intro {
  font-size: 14px;
  color: var(--ink-soft);
  margin-bottom: 18px;
  line-height: 1.6;
}

/* === HERO CARD === */
.hero-card {
  background: linear-gradient(135deg, var(--teal) 0%, var(--teal-mid) 100%);
  color: white;
  border-radius: var(--radius);
  padding: 36px 32px;
  margin-bottom: 12px;
  box-shadow: var(--shadow-lg);
  position: relative;
  overflow: hidden;
}
.hero-card::before {
  content: '';
  position: absolute;
  top: -40px; right: -40px;
  width: 200px; height: 200px;
  background: radial-gradient(circle, rgba(255,255,255,0.1), transparent 70%);
  border-radius: 50%;
}
.hero-stufe {
  display: inline-block;
  font-size: 12px;
  font-weight: 600;
  text-transform: uppercase;
  letter-spacing: 0.12em;
  opacity: 0.85;
  padding: 5px 12px;
  background: rgba(255,255,255,0.15);
  border-radius: 999px;
  margin-bottom: 20px;
}
.hero-gehalt {
  font-size: 56px;
  font-weight: 800;
  letter-spacing: -0.03em;
  line-height: 1;
  margin-bottom: 4px;
}
.hero-gehalt-unit {
  font-size: 18px;
  font-weight: 500;
  opacity: 0.9;
  margin-bottom: 20px;
}
.hero-meta {
  display: flex; gap: 24px; flex-wrap: wrap;
  padding-top: 20px;
  border-top: 1px solid rgba(255,255,255,0.15);
  font-size: 13px;
}
.hero-meta-item { opacity: 0.9; }
.hero-meta-label {
  font-size: 11px;
  text-transform: uppercase;
  letter-spacing: 0.08em;
  opacity: 0.7;
  display: block;
  margin-bottom: 2px;
}
.hero-meta-value { font-weight: 700; font-size: 15px; }

/* === BREAKDOWN === */
.breakdown-card {
  background: var(--white);
  border-radius: var(--radius);
  padding: 24px 28px;
  box-shadow: var(--shadow-sm);
  border: 1px solid var(--line);
}
.breakdown-formula {
  font-size: 15px;
  color: var(--ink-soft);
  margin-bottom: 18px;
  line-height: 1.7;
}
.breakdown-bar {
  display: flex; height: 40px;
  border-radius: 10px;
  overflow: hidden;
  margin-bottom: 12px;
  box-shadow: inset 0 0 0 1px var(--line);
}
.breakdown-seg {
  display: flex; align-items: center; justify-content: center;
  font-size: 12px; font-weight: 700; color: white;
  white-space: nowrap;
  padding: 0 10px;
}
.breakdown-seg.sockel { background: var(--teal); }
.breakdown-seg.bundle { background: var(--teal-mid); }
.breakdown-seg.stufe { background: var(--orange); }
.breakdown-seg.rest { background: var(--line); color: var(--muted); }

.breakdown-legend {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
  gap: 12px;
  margin-top: 14px;
}
.breakdown-legend-item {
  display: flex; flex-direction: column; gap: 2px;
  padding-left: 12px;
  border-left: 3px solid var(--teal);
}
.breakdown-legend-item.bundle { border-color: var(--teal-mid); }
.breakdown-legend-item.stufe { border-color: var(--orange); }
.breakdown-legend-label {
  font-size: 11px; color: var(--muted);
  text-transform: uppercase; letter-spacing: 0.05em;
}
.breakdown-legend-val { font-size: 16px; font-weight: 700; color: var(--ink); }
.breakdown-result {
  margin-top: 20px;
  padding-top: 20px;
  border-top: 1px dashed var(--line-strong);
  display: flex; align-items: baseline; justify-content: space-between;
  flex-wrap: wrap; gap: 12px;
}
.breakdown-result-label { font-size: 13px; color: var(--muted); }
.breakdown-result-val { font-size: 28px; font-weight: 800; color: var(--teal); }

/* === TWO-KPI GRID === */
.kpi-grid {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 14px;
}
.kpi-card {
  background: var(--white);
  border-radius: var(--radius-sm);
  padding: 22px 20px;
  border: 1px solid var(--line);
  box-shadow: var(--shadow-sm);
}
.kpi-label {
  font-size: 11px;
  color: var(--muted);
  text-transform: uppercase;
  letter-spacing: 0.08em;
  font-weight: 600;
  margin-bottom: 8px;
}
.kpi-value {
  font-size: 34px; font-weight: 800; letter-spacing: -0.02em;
  line-height: 1; margin-bottom: 4px;
}
.kpi-value.good { color: var(--green); }
.kpi-value.ok { color: var(--teal); }
.kpi-value.warn { color: var(--orange); }
.kpi-unit { font-size: 14px; font-weight: 500; color: var(--muted); margin-left: 2px; }
.kpi-schwelle {
  font-size: 12px; color: var(--muted); margin-top: 10px;
  padding-top: 10px; border-top: 1px solid var(--line);
}
.kpi-schwelle strong { color: var(--ink-soft); font-weight: 600; }
.kpi-bar {
  margin-top: 12px;
  height: 6px; background: var(--line); border-radius: 3px; overflow: hidden;
}
.kpi-bar-fill {
  height: 100%; border-radius: 3px;
  background: linear-gradient(90deg, var(--teal), var(--teal-mid));
  transition: width 0.8s ease;
}
.kpi-bar-fill.over { background: linear-gradient(90deg, var(--green), #66bb6a); }
.kpi-bar-fill.under { background: linear-gradient(90deg, var(--orange), #ffab40); }

/* === STUFEN-LEITER === */
.stufen-scroll {
  display: flex;
  gap: 10px;
  overflow-x: auto;
  padding: 8px 2px 14px;
  scrollbar-width: thin;
  scrollbar-color: var(--line-strong) transparent;
}
.stufen-scroll::-webkit-scrollbar { height: 4px; }
.stufen-scroll::-webkit-scrollbar-thumb { background: var(--line-strong); border-radius: 2px; }

.stufe-chip {
  flex: 0 0 auto;
  background: var(--white);
  border: 2px solid var(--line);
  border-radius: 14px;
  padding: 14px 16px;
  min-width: 120px;
  text-align: center;
  transition: all 0.2s;
}
.stufe-chip-num {
  font-size: 10px; font-weight: 700;
  color: var(--muted); letter-spacing: 0.1em;
  text-transform: uppercase;
  margin-bottom: 2px;
}
.stufe-chip-name {
  font-size: 14px; font-weight: 700; color: var(--ink);
  margin-bottom: 8px;
}
.stufe-chip-detail {
  font-size: 11px; color: var(--muted); line-height: 1.4;
}
.stufe-chip.current {
  border-color: var(--teal); background: var(--teal-light);
  transform: scale(1.05);
  box-shadow: var(--shadow);
}
.stufe-chip.current .stufe-chip-name { color: var(--teal); }
.stufe-chip.next {
  border-color: var(--orange);
  border-style: dashed;
  background: var(--orange-light);
}
.stufe-chip.next .stufe-chip-name { color: var(--orange); }

/* === GAP BLOCK === */
.gap-card {
  background: linear-gradient(135deg, var(--orange-light) 0%, #fff5e6 100%);
  border-radius: var(--radius);
  padding: 24px 28px;
  border: 1px solid #ffd9b3;
}
.gap-header {
  display: flex; align-items: center; gap: 10px;
  margin-bottom: 16px;
}
.gap-from, .gap-to {
  display: inline-flex; align-items: center;
  padding: 4px 12px; border-radius: 999px;
  font-size: 12px; font-weight: 700;
}
.gap-from { background: var(--teal-light); color: var(--teal); }
.gap-to { background: var(--orange-light); color: var(--orange); border: 1px solid #ffb27a; }
.gap-arrow { color: var(--muted); font-size: 18px; }
.gap-numbers {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 16px;
  margin-bottom: 16px;
}
.gap-number { }
.gap-number-label {
  font-size: 11px; color: var(--muted);
  text-transform: uppercase; letter-spacing: 0.08em;
  font-weight: 600;
  margin-bottom: 4px;
}
.gap-number-val { font-size: 22px; font-weight: 800; color: var(--orange); }
.gap-number-val.pos { color: var(--green); }
.gap-bar-wrap {
  background: rgba(255,255,255,0.6);
  border-radius: 8px;
  padding: 10px 14px;
}
.gap-bar-label {
  font-size: 11px; color: var(--muted);
  text-transform: uppercase; letter-spacing: 0.08em;
  font-weight: 600;
  margin-bottom: 6px;
}
.gap-bar {
  height: 10px; background: white; border-radius: 5px;
  overflow: hidden; margin-bottom: 6px;
  border: 1px solid #ffd9b3;
}
.gap-bar-fill {
  height: 100%; border-radius: 5px;
  background: linear-gradient(90deg, var(--orange), #ffab40);
  transition: width 0.8s ease;
}
.gap-bar-caption { font-size: 11px; color: var(--muted); text-align: right; }

/* === HEBEL === */
.hebel-grid {
  display: grid; gap: 10px;
}
.hebel-item {
  background: var(--white);
  border-radius: var(--radius-sm);
  padding: 16px 18px;
  border: 1px solid var(--line);
  display: grid;
  grid-template-columns: 1fr auto;
  gap: 16px;
  align-items: center;
}
.hebel-content { }
.hebel-name { font-weight: 700; color: var(--ink); margin-bottom: 4px; }
.hebel-desc { font-size: 13px; color: var(--ink-soft); line-height: 1.5; }
.hebel-effect {
  font-size: 12px; font-weight: 700;
  color: var(--teal);
  background: var(--teal-light);
  padding: 8px 12px;
  border-radius: 8px;
  white-space: nowrap;
  text-align: center;
}
.hebel-effect.borderline { color: var(--orange); background: var(--orange-light); }
.hebel-effect.impossible { color: var(--red); background: var(--red-light); }
.hebel-tag {
  display: inline-block;
  font-size: 11px; font-weight: 600;
  padding: 2px 8px; border-radius: 999px;
  margin-left: 6px;
  letter-spacing: 0.02em;
}
.hebel-tag.realistic  { background: var(--green-light);  color: var(--green); }
.hebel-tag.borderline { background: var(--orange-light); color: var(--orange); }
.hebel-tag.impossible { background: var(--red-light);    color: var(--red); }
.hebel-from-to {
  font-size: 12px; color: var(--ink-soft);
  margin-top: 4px;
}
.hebel-headline {
  font-size: 15px; line-height: 1.55;
  margin-bottom: 14px; color: var(--ink-soft);
}
.hebel-headline strong { color: var(--ink); }
.hebel-note {
  font-size: 12px; color: var(--muted);
  margin-top: 10px; font-style: italic;
}

/* === TIMELINE === */
.timeline-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(120px, 1fr));
  gap: 10px;
}
.timeline-item {
  background: var(--white);
  border: 1px solid var(--line);
  border-radius: var(--radius-sm);
  padding: 14px 16px;
}
.timeline-item.current {
  border-color: var(--teal); background: var(--teal-light);
}
.timeline-item.empty {
  background: #f7f7f7; border-style: dashed;
}
.timeline-q {
  font-size: 11px; color: var(--muted); font-weight: 600;
  text-transform: uppercase; letter-spacing: 0.06em;
  margin-bottom: 6px;
}
.timeline-stufe { font-size: 14px; font-weight: 700; color: var(--ink); margin-bottom: 2px; }
.timeline-gehalt { font-size: 13px; color: var(--ink-soft); }
.timeline-empty { font-size: 12px; color: var(--muted); font-style: italic; }

/* === LIVE === */
.live-card {
  background: var(--white);
  border: 2px solid var(--teal-mid);
  border-radius: var(--radius);
  padding: 24px 28px;
  box-shadow: var(--shadow);
}
.live-status {
  display: inline-flex; align-items: center; gap: 8px;
  font-size: 12px; font-weight: 700;
  color: var(--teal);
  padding: 4px 12px;
  background: var(--teal-light);
  border-radius: 999px;
  margin-bottom: 16px;
}
.live-status::before {
  content: ''; display: inline-block;
  width: 8px; height: 8px; border-radius: 50%;
  background: var(--green); animation: pulse 2s infinite;
}
@keyframes pulse {
  0%, 100% { opacity: 1; transform: scale(1); }
  50% { opacity: 0.5; transform: scale(1.3); }
}
.live-placeholder {
  font-size: 14px; color: var(--muted); line-height: 1.6;
  padding: 16px;
  background: var(--bg);
  border-radius: var(--radius-sm);
}

/* === FOOTER === */
.footer {
  margin-top: 48px; padding-top: 24px;
  border-top: 1px solid var(--line);
  text-align: center;
  font-size: 12px; color: var(--muted);
}
.footer a { color: var(--teal); text-decoration: none; }


.gap-row {
  display: flex;
  gap: 14px;
  padding: 18px 0;
  border-top: 1px solid #ffd9b3;
}
.gap-row:first-child { border-top: none; padding-top: 4px; }
.gap-row:last-child { padding-bottom: 4px; }
.gap-row-icon {
  width: 32px; height: 32px; border-radius: 50%;
  display: flex; align-items: center; justify-content: center;
  font-size: 16px; font-weight: 700; flex-shrink: 0;
}
.gap-row.done .gap-row-icon {
  background: var(--green-light); color: var(--green);
}
.gap-row.gap .gap-row-icon {
  background: var(--orange-light); color: var(--orange);
}
.gap-row-content { flex: 1; min-width: 0; }
.gap-row-title {
  font-weight: 700; font-size: 15px; color: var(--ink);
  margin-bottom: 8px;
}
.gap-row.done .gap-row-title { color: var(--green); }
.gap-row-detail {
  font-size: 13px; color: var(--ink-soft); line-height: 1.5;
}
.gap-row-kpis {
  display: grid; grid-template-columns: auto 1fr; gap: 12px 18px;
  margin-bottom: 10px;
}
.gap-mini-label {
  font-size: 11px; color: var(--muted);
  text-transform: uppercase; letter-spacing: 0.05em;
  margin-bottom: 2px;
}
.gap-mini-val { font-size: 17px; font-weight: 700; color: var(--orange); }

@media (max-width: 600px) {
  .gap-row-kpis { grid-template-columns: 1fr; }
}

/* === ZUFRIEDENHEITS-BREAKDOWN === */
.zufr-breakdown {
  margin-top: 14px; padding-top: 14px;
  border-top: 1px dashed var(--line);
}
.zufr-breakdown-title {
  font-size: 10px; font-weight: 700; color: var(--muted);
  text-transform: uppercase; letter-spacing: 0.08em;
  margin-bottom: 10px;
}
.zufr-sub {
  display: grid; grid-template-columns: 1fr auto; gap: 4px 10px;
  align-items: center;
  margin-bottom: 8px;
}
.zufr-sub:last-child { margin-bottom: 0; }
.zufr-sub-label {
  font-size: 12px; color: var(--ink-soft);
  grid-column: 1 / 3;
}
.zufr-sub-bar {
  grid-column: 1;
  height: 5px; background: var(--line); border-radius: 3px;
  overflow: hidden;
}
.zufr-sub-fill {
  display: block; height: 100%;
  background: linear-gradient(90deg, var(--teal-mid), var(--teal));
  border-radius: 3px;
}
.zufr-sub-val {
  grid-column: 2; grid-row: 2;
  font-size: 13px; font-weight: 700; color: var(--teal);
  white-space: nowrap;
}
.zufr-sub-val small { font-weight: 500; color: var(--muted); margin-left: 2px; font-size: 11px; }

/* === LIVE KPIs === */
.live-kpi-card {
  background: linear-gradient(135deg, #fff 0%, var(--bg) 100%);
  border: 2px solid var(--teal-light);
  border-radius: var(--radius);
  padding: 20px 22px;
  margin-bottom: 18px;
  box-shadow: var(--shadow-sm);
}
.live-kpi-header {
  display: flex; justify-content: space-between; align-items: center;
  margin-bottom: 14px; padding-bottom: 12px;
  border-bottom: 1px dashed var(--line);
}
.live-kpi-title {
  font-size: 14px; font-weight: 800; color: var(--teal);
  text-transform: uppercase; letter-spacing: 0.05em;
}
.live-kpi-date {
  font-size: 11px; color: var(--muted);
  background: var(--teal-light);
  padding: 3px 10px; border-radius: 999px;
  color: var(--teal); font-weight: 600;
}
.live-kpi-row {
  display: grid;
  grid-template-columns: 130px 1fr;
  gap: 10px 14px;
  align-items: center;
  margin-bottom: 12px;
}
.live-kpi-row:last-child { margin-bottom: 0; }
.live-kpi-label {
  font-size: 11px; font-weight: 700;
  text-transform: uppercase; letter-spacing: 0.08em;
  color: var(--muted);
}
.live-kpi-chip-track {
  position: relative;
  background: var(--bg);
  border-radius: 10px;
  height: 36px;
  overflow: hidden;
  border: 1px solid var(--line);
}
.live-kpi-chip {
  position: absolute; inset: 0 auto 0 0;
  display: flex; justify-content: space-between; align-items: center;
  padding: 0 12px;
  border-radius: 10px;
  font-size: 13px; font-weight: 700;
  color: white;
  box-shadow: var(--shadow-sm);
  transition: width 0.6s ease;
}
.live-kpi-note {
  grid-column: 2 / 3;
  font-size: 11px; color: var(--muted);
  font-style: italic;
  margin-top: -4px;
}
@media (max-width: 600px) {
  .live-kpi-row { grid-template-columns: 1fr; gap: 4px; }
  .live-kpi-note { grid-column: 1; }
}

/* === WEGE === */
.wege-grid { display: grid; gap: 14px; }
.weg-card {
  background: var(--white);
  border: 1px solid var(--line);
  border-radius: var(--radius);
  padding: 20px 22px;
  box-shadow: var(--shadow-sm);
  position: relative;
  overflow: hidden;
}
.weg-card::before {
  content: ''; position: absolute;
  top: 0; left: 0; width: 4px; height: 100%;
  background: linear-gradient(180deg, var(--teal), var(--teal-mid));
}
.weg-head {
  display: flex; align-items: center; justify-content: space-between;
  margin-bottom: 16px; padding-bottom: 14px;
  border-bottom: 1px dashed var(--line);
}
.weg-head-label {
  font-size: 17px; font-weight: 800; color: var(--teal);
  letter-spacing: -0.01em;
}
.weg-head-tag {
  font-size: 10px; font-weight: 700;
  text-transform: uppercase; letter-spacing: 0.1em;
  color: var(--muted);
  background: var(--bg); padding: 4px 10px; border-radius: 999px;
}
.weg-vars { display: grid; gap: 10px; }
.weg-var {
  display: grid; grid-template-columns: 110px 1fr;
  align-items: center; gap: 14px;
}
.weg-var-label {
  font-size: 11px; font-weight: 700;
  text-transform: uppercase; letter-spacing: 0.08em;
  color: var(--muted);
}
.weg-chip-track {
  position: relative;
  background: var(--bg);
  border-radius: 10px;
  height: 40px;
  overflow: hidden;
  border: 1px solid var(--line);
}
.weg-chip {
  position: absolute; inset: 0 auto 0 0;
  display: flex; justify-content: space-between; align-items: center;
  padding: 0 14px;
  border-radius: 10px;
  font-size: 14px; font-weight: 700;
  color: white;
  box-shadow: var(--shadow-sm);
  transition: width 0.6s ease;
}
.weg-chip-text { font-size: 14px; font-weight: 700; white-space: nowrap; }
.weg-chip-range {
  font-size: 12px; font-weight: 600;
  background: rgba(255,255,255,0.22);
  padding: 2px 8px; border-radius: 999px;
  white-space: nowrap;
}
.weg-chip-low   { width: 32%; background: linear-gradient(135deg, #8bc34a, #7cb342); }
.weg-chip-mid   { width: 55%; background: linear-gradient(135deg, #ffb547, #ff9800); }
.weg-chip-high  { width: 78%; background: linear-gradient(135deg, var(--teal-mid), var(--teal)); }
.weg-chip-vhigh { width: 100%; background: linear-gradient(135deg, #095859, #054546); }

.weg-text {
  background: var(--bg); border-radius: var(--radius-sm);
  padding: 20px 22px; color: var(--ink-soft);
  font-size: 14px; line-height: 1.6;
  border-left: 3px solid var(--teal);
}

.weg-legende {
  margin-top: 18px; padding: 16px 18px;
  background: var(--bg); border-radius: var(--radius-sm);
  font-size: 12px; color: var(--muted);
  border: 1px solid var(--line);
}
.weg-legende-title {
  font-weight: 700; font-size: 11px;
  text-transform: uppercase; letter-spacing: 0.08em;
  color: var(--ink-soft); margin-bottom: 12px;
}
.weg-legende-row {
  display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 10px; align-items: center;
}
.weg-legende-row:last-child { margin-bottom: 0; }
.weg-legende-kpi {
  font-weight: 700; color: var(--ink-soft); font-size: 12px;
  min-width: 90px;
}
.weg-legende-item {
  display: inline-flex; align-items: center; gap: 6px;
  padding: 3px 10px; border-radius: 999px;
  color: white; font-size: 11px; font-weight: 700;
}
.weg-legende-item.low   { background: #8bc34a; }
.weg-legende-item.mid   { background: #ff9800; }
.weg-legende-item.high  { background: var(--teal); }
.weg-legende-item.vhigh { background: #095859; }
.weg-legende-item .rg {
  opacity: 0.85; font-weight: 500; font-size: 10px;
}

/* === RESPONSIVE === */
@media (max-width: 600px) {
  .container { padding: 20px 16px 60px; }
  .hero-gehalt { font-size: 44px; }
  .kpi-grid { grid-template-columns: 1fr; }
  .gap-numbers { grid-template-columns: 1fr; }
  .breakdown-legend { grid-template-columns: 1fr 1fr; }
  .hebel-item { grid-template-columns: 1fr; gap: 10px; }
  .hebel-effect { width: 100%; }
}
"""

def fmt_eur(n, decimals=0):
    if n is None: return '—'
    return f"{n:,.{decimals}f}".replace(',', '§').replace('.', ',').replace('§', '.')

def fmt_de(n, decimals=1):
    """Deutsche Dezimal-Formatierung (Komma statt Punkt) für Zahlen ohne €."""
    if n is None: return '—'
    return f"{n:,.{decimals}f}".replace(',', '§').replace('.', ',').replace('§', '.')

def quartal_label(d):
    """'Q2 2026' für ein date-Objekt."""
    return f"Q{(d.month - 1) // 3 + 1} {d.year}"

def vorquartal_label(d):
    """Letztes abgeschlossenes Quartal vor d, z.B. 'Q1 2026' wenn d in Q2 2026."""
    q = (d.month - 1) // 3 + 1
    return f"Q{q - 1} {d.year}" if q > 1 else f"Q4 {d.year - 1}"

def _kpi_bar_render(val, curr_threshold, next_threshold, curr_n, next_n, has_next):
    """Bar mit klar getrennter Caption: '✓ Stufe X erreicht · Y % Richtung Stufe Z'.
    Bar-Füllung = Progress zwischen aktueller und nächster Schwelle (0–100 %)."""
    if not has_next:
        cap = '<span style="color:var(--green);font-weight:700;">✓ Höchste Stufe erreicht</span>'
        return f'<div class="kpi-bar-caption" style="font-size:11px;margin-top:8px;">{cap}</div><div class="kpi-bar"><div class="kpi-bar-fill over" style="width:100%"></div></div>'
    if val >= next_threshold:
        cap = f'<span style="color:var(--green);font-weight:700;">✓ über Schwelle Stufe {next_n}</span>'
        return f'<div class="kpi-bar-caption" style="font-size:11px;margin-top:8px;">{cap}</div><div class="kpi-bar"><div class="kpi-bar-fill over" style="width:100%"></div></div>'
    if val >= curr_threshold:
        denom = next_threshold - curr_threshold
        progress = ((val - curr_threshold) / denom * 100) if denom > 0 else 100
        cap = f'<span style="color:var(--green);font-weight:700;">✓ Stufe {curr_n} erreicht</span> · {progress:.0f} % Richtung Stufe {next_n}'
        return f'<div class="kpi-bar-caption" style="font-size:11px;margin-top:8px;color:var(--ink-soft);">{cap}</div><div class="kpi-bar"><div class="kpi-bar-fill" style="width:{progress:.0f}%"></div></div>'
    cap = f'<span style="color:var(--orange);font-weight:700;">✗ Schwelle Stufe {curr_n} nicht erreicht</span>'
    return f'<div class="kpi-bar-caption" style="font-size:11px;margin-top:8px;">{cap}</div><div class="kpi-bar"><div class="kpi-bar-fill under" style="width:0%"></div></div>'

def _eur_bar_html(pm, curr_s, next_s, d):
    return _kpi_bar_render(pm['eur60'], curr_s['eur60'], next_s['eur60'], curr_s['n'], next_s['n'], bool(d))

def _zufr_bar_html(pm, curr_s, next_s, d):
    return _kpi_bar_render(pm['zufr'], curr_s['zufr'], next_s['zufr'], curr_s['n'], next_s['n'], bool(d))



# ============ LIVE KPIs ============

AUSLASTUNG_4W_TABLE = 'm29vw64nhicfco2'
TH_BY_NAME_CACHE = {}

import subprocess, tempfile, os as _os, time as _time

def _env():
    if _os.environ.get('NOCODB_TOKEN'):
        return {'NOCODB_TOKEN': _os.environ['NOCODB_TOKEN']}
    e = {}
    env_path = _os.path.expanduser('~/.claude/.env')
    if not _os.path.exists(env_path):
        raise RuntimeError('NOCODB_TOKEN not set and ~/.claude/.env not found')
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if '=' in line and not line.startswith('#'):
                k, v = line.split('=', 1)
                e[k] = v.strip('"').strip("'")
    return e

_NOCO_CACHE = {}

def _fetch_all(table_id, where=None):
    cache_key = f'{table_id}|{where}'
    if cache_key in _NOCO_CACHE:
        return _NOCO_CACHE[cache_key]
    env = _env()
    rows = []; offset = 0
    cfg = tempfile.NamedTemporaryFile(mode='w', suffix='.conf', delete=False)
    cfg.write(f'header = "xc-token: {env["NOCODB_TOKEN"]}"\n'); cfg.close(); _os.chmod(cfg.name, 0o600)
    try:
        while True:
            params = [f'limit=200', f'offset={offset}']
            if where: params.append(f'where={where}')
            url = f'https://db.vacura-praxis.de/api/v2/tables/{table_id}/records?' + '&'.join(params)
            # Retry 5x mit Exponential Backoff (1s, 2s, 4s, 8s, 16s)
            data = None
            last_err = None
            for attempt in range(5):
                try:
                    r = subprocess.run(['curl','-sS','--max-time','60','-K',cfg.name,url], capture_output=True, text=True, timeout=90)
                    if r.returncode != 0 or not r.stdout.strip():
                        last_err = f'curl rc={r.returncode}, stderr={r.stderr[:200]}'
                        _time.sleep(2 ** attempt); continue
                    data = json.loads(r.stdout)
                    break
                except Exception as e:
                    last_err = str(e)[:200]
                    _time.sleep(2 ** attempt)
            if not data:
                raise RuntimeError(f'NocoDB fetch failed after 5 retries for {table_id}: {last_err}')
            batch = data.get('list', [])
            rows.extend(batch)
            if len(batch) < 200:
                _NOCO_CACHE[cache_key] = rows
                return rows
            offset += 200
    finally:
        _os.unlink(cfg.name)

# === Vergütungssätze (NocoDB `verguetungssaetze`, id mi2h0wengv4xbjh) ===
# Nur Kategorien Dauer-Basis / Aufschlag / Zuschlag — Versicherungsfaktoren bleiben
# hardcoded (PKV_FAKTOR oben). Datum-aware: pro Termin wird der zum `beginn` gültige
# Tarif gewählt (gueltig_ab ≤ Datum, gueltig_bis NULL oder ≥ Datum, bei mehreren das
# jüngste gueltig_ab).
_VERG_TABLE_ID = 'mi2h0wengv4xbjh'

# Fallback Stand 2026-06 — wird nur genutzt wenn NocoDB nicht erreichbar ist oder
# einzelne Schlüssel fehlen. Tests pinnen den Cache via conftest.py auf diese Liste,
# damit die Test-Suite deterministisch + offline läuft.
_TARIFE_FALLBACK = [
    {'schluessel': 'basis_bis_20',                'wert': 8.51,  'gueltig_ab': '2024-01-01', 'gueltig_bis': None},
    {'schluessel': 'basis_bis_30',                'wert': 56.93, 'gueltig_ab': '2024-01-01', 'gueltig_bis': None},
    {'schluessel': 'basis_bis_45',                'wert': 75.91, 'gueltig_ab': '2024-01-01', 'gueltig_bis': None},
    {'schluessel': 'basis_bis_60',                'wert': 94.89, 'gueltig_ab': '2024-01-01', 'gueltig_bis': None},
    {'schluessel': 'aufschlag_je_15min_ueber_60', 'wert': 18.98, 'gueltig_ab': '2024-01-01', 'gueltig_bis': None},
    {'schluessel': 'hausbesuch_pauschale',        'wert': 27.56, 'gueltig_ab': '2024-01-01', 'gueltig_bis': None},
]

_TARIFE_CACHE = None

def _load_tarife():
    global _TARIFE_CACHE
    if _TARIFE_CACHE is not None:
        return _TARIFE_CACHE
    try:
        rows = _fetch_all(_VERG_TABLE_ID, where='(kategorie,in,Dauer-Basis,Aufschlag,Zuschlag)')
        if not rows:
            raise RuntimeError('verguetungssaetze: leere Rückgabe')
        _TARIFE_CACHE = rows
        # Diagnostik zu stderr, damit --check-tarife JSON-only auf stdout bleibt
        print(f'  Tarife aus NocoDB: {len(rows)} Vergütungssätze (Dauer-Basis/Aufschlag/Zuschlag)', file=sys.stderr)
    except Exception as e:
        print(f'  (Tarife-Fetch fehlgeschlagen: {str(e)[:120]} — Fallback auf hardcoded Stand 2026-06)', file=sys.stderr)
        _TARIFE_CACHE = _TARIFE_FALLBACK
    return _TARIFE_CACHE

def _tarif_for(schluessel, datum_iso10):
    """Gültigen Tarif zum Datum (YYYY-MM-DD) liefern. Bei mehreren passenden das jüngste
    `gueltig_ab`. Bei fehlendem Schlüssel Fallback auf _TARIFE_FALLBACK."""
    cands = [r for r in _load_tarife() if r.get('schluessel') == schluessel]
    valid = [r for r in cands
             if (r.get('gueltig_ab') or '0000-00-00')[:10] <= datum_iso10
             and (not r.get('gueltig_bis') or r['gueltig_bis'][:10] >= datum_iso10)]
    if not valid:
        for r in _TARIFE_FALLBACK:
            if r['schluessel'] == schluessel:
                return float(r['wert'])
        raise KeyError(f'Tarif-Schluessel "{schluessel}" weder in NocoDB noch im Fallback')
    return float(max(valid, key=lambda r: r.get('gueltig_ab') or '')['wert'])

# Default-Termin-Mix für Skalierungsfaktor bei Tarif-Wechsel (Vacura-typisch, grob kalibriert
# Stand 2026-06; bei größeren Mix-Verschiebungen empirisch nachsampeln). Summe ≈ 1.
DEFAULT_TERMIN_MIX = {
    'basis_bis_20': 0.01,   # TA selten
    'basis_bis_30': 0.50,   # Standard-30-min
    'basis_bis_45': 0.39,   # 45-min Hauptanteil
    'basis_bis_60': 0.10,
}
DEFAULT_HB_ANTEIL = 0.08    # Anteil HB-Termine (Pauschale wird additiv mit-skaliert)

def _check_tarif_aenderungen(today_d, lookback_days=7, mix=None, hb_anteil=None):
    """Prüft, ob neue Vergütungssätze in den letzten X Tagen aktiv geworden sind.

    Vergleicht jeden Schlüssel mit dem unmittelbaren Vorgänger (höchstes
    `gueltig_ab` vor dem neuen) und berechnet einen gewichteten Skalierungsfaktor
    aus DEFAULT_TERMIN_MIX + DEFAULT_HB_ANTEIL. Liefert empfohlene neue
    Stufentabellen-Schwellen (aktuelle STUFEN × Faktor) — die Übernahme bleibt
    Geschäftsführungs-Entscheidung (siehe tarif-watch.yml Issue-Body).

    Returns dict mit 'changes', 'scale_factor', 'empfohlen_schwellen' — oder None
    falls keine Änderung im Fenster."""
    from datetime import timedelta as _td, date as _date
    if isinstance(today_d, str):
        today_d = _date.fromisoformat(today_d[:10])
    today_iso = today_d.isoformat()
    cutoff = (today_d - _td(days=lookback_days)).isoformat()

    tarife = _load_tarife()
    new_rows = [r for r in tarife
                if cutoff <= (r.get('gueltig_ab') or '')[:10] <= today_iso]
    if not new_rows:
        return None

    changes = []
    for new in new_rows:
        s = new.get('schluessel')
        if not s:
            continue
        old_rows = [r for r in tarife
                    if r.get('schluessel') == s
                    and (r.get('gueltig_ab') or '')[:10] < (new.get('gueltig_ab') or '')[:10]]
        if not old_rows:
            continue
        alt = max(old_rows, key=lambda r: r.get('gueltig_ab') or '')
        alt_w = float(alt['wert']); neu_w = float(new['wert'])
        changes.append({
            'schluessel': s,
            'alt_wert': alt_w,
            'neu_wert': neu_w,
            'gueltig_ab': new['gueltig_ab'][:10],
            'delta_pct': round((neu_w / alt_w - 1) * 100, 2) if alt_w else 0,
        })

    if not changes:
        return None

    mix = mix or DEFAULT_TERMIN_MIX
    hb_anteil = DEFAULT_HB_ANTEIL if hb_anteil is None else hb_anteil
    chg_by_key = {c['schluessel']: c for c in changes}

    def _val(schl, side):
        if schl in chg_by_key:
            return chg_by_key[schl][f'{side}_wert']
        for r in _TARIFE_FALLBACK:
            if r['schluessel'] == schl:
                return float(r['wert'])
        return 0.0

    sum_alt = sum(w * _val(k, 'alt') for k, w in mix.items())
    sum_neu = sum(w * _val(k, 'neu') for k, w in mix.items())
    sum_alt += hb_anteil * _val('hausbesuch_pauschale', 'alt')
    sum_neu += hb_anteil * _val('hausbesuch_pauschale', 'neu')
    scale = (sum_neu / sum_alt) if sum_alt else 1.0

    empfohlen = [{
        'n': s['n'], 'name': s['name'],
        'alt_eur60': round(s['eur60'], 2),
        'neu_eur60': round(s['eur60'] * scale, 2),
    } for s in STUFEN]

    return {
        'changes': changes,
        'scale_factor': round(scale, 4),
        'empfohlen_schwellen': empfohlen,
        'mix_used': mix,
        'hb_anteil': hb_anteil,
    }

def termin_umsatz(t):
    """€-Umsatz eines Termins nach ZI-Systematik (siehe Konstanten-Block oben).

    (round(Dauer/15) + 1 VNB-ZI) × 18,98 € — die Kalenderdauer ist reine Behandlungszeit,
    die Vor-/Nachbereitungs-ZI wird obendrauf abgerechnet. Thermische Anwendung/KT/WT
    pauschal 8,51 €. PKV ×2,0, Selbstzahler ×1,7, Hausbesuch +27,56 € (Pauschale NICHT
    × Faktor — Entscheidung 2026-06-03). +4,11 % ab 01.07.2026.
    Backtest vs. MediFox-Bundle-Werte Q1+Q2 2026: −4,4 bis +1,7 % (inkl. VO-Gebühren)."""
    from datetime import datetime as _dt
    try:
        beginn = _dt.fromisoformat(t['beginn'].replace('Z', '+00:00'))
        ende   = _dt.fromisoformat(t['ende'].replace('Z', '+00:00'))
        dauer  = (ende - beginn).total_seconds() / 60
    except Exception:
        return 0.0
    if dauer <= 0: return 0.0
    datum = beginn.date().isoformat()
    f = satz_faktor(datum)
    bez = str(t.get('bezeichnung') or '').lower()
    if 'thermisch' in bez or 'kälte' in bez or 'wärme' in bez or bez.strip() in ('wt', 'kt', 'urb'):
        basis = THERMISCH_PREIS * f
    else:
        zi = round(dauer / 15) + 1
        basis = zi * ZI_PREIS * f
    if t.get('verordnungstyp') == 2:
        basis *= PKV_FAKTOR
    elif t.get('verordnungstyp') == 3:
        basis *= SZ_FAKTOR
    if t.get('is_hausbesuch'):
        basis += HB_PAUSCHALE * f
    return basis

BERLIN_FEIERTAGE = {
    # 2026
    '2026-01-01',  # Neujahr (Do)
    '2026-03-08',  # Internationaler Frauentag (So 2026 — nur in Berlin gesetzlich)
    '2026-04-03',  # Karfreitag (Fr)
    '2026-04-06',  # Ostermontag (Mo)
    '2026-05-01',  # Tag der Arbeit (Fr)
    '2026-05-14',  # Christi Himmelfahrt (Do)
    '2026-05-25',  # Pfingstmontag (Mo)
    '2026-10-03',  # Tag der Deutschen Einheit (Sa)
    '2026-12-25',  # 1. Weihnachtstag (Fr)
    '2026-12-26',  # 2. Weihnachtstag (Sa)
    # 2027
    '2027-01-01', '2027-03-08', '2027-03-26', '2027-03-29',
    '2027-05-01', '2027-05-06', '2027-05-17',
    '2027-10-03', '2027-12-25', '2027-12-26',
}

def is_probezeit(startdatum, q_eval_end):
    """True wenn PM am q_eval_end weniger als 6 Monate beschäftigt war (§ 8 Abs. 3 v9-Vertrag).

    Gemeinsame Helper-Funktion für compute_pm und compute_quartal — eine einzige
    Wahrheit für die Probezeit-Regel, damit beide Pfade niemals auseinanderdriften.
    """
    if not startdatum: return False
    try:
        from datetime import date as _date
        sd = startdatum.date() if hasattr(startdatum, 'date') else (
            _date.fromisoformat(startdatum[:10]) if isinstance(startdatum, str) else startdatum)
        monatsdiff = (q_eval_end.year - sd.year) * 12 + (q_eval_end.month - sd.month)
        return monatsdiff < 6
    except Exception:
        return False


def _th_earliest_beschaeftigung(m):
    """Frühestes Beschäftigungs-Von-Datum (ISO) für 29-Tage-Sperre."""
    bz = m.get('beschaeftigungszeiten') or []
    starts = [e['Von'] for e in bz if e.get('Von')]
    return min(starts) if starts else None

def _anlauf_cutoff(today, months=3):
    """Stichtag: TH mit Start NACH diesem Datum sind in der Anlaufphase und zählen noch
    nicht im 4W-Schnitt (= n8n minMonateAktiv=3, include_in_site=false). NocoDB hat das
    Flag nicht, daher rekonstruieren wir die Schwelle über das Beschäftigungs-Startdatum."""
    from calendar import monthrange
    m = today.month - months - 1
    y = today.year + m // 12
    m = m % 12 + 1
    return date(y, m, min(today.day, monthrange(y, m)[1]))

def _ist_anlauf_th(m, cutoff):
    """True, wenn der TH-Start nach dem Anlauf-Cutoff liegt (noch nicht im Schnitt)."""
    s = _th_earliest_beschaeftigung(m)
    if not s:
        return False
    try:
        return date.fromisoformat(s[:10]) > cutoff
    except Exception:
        return False

def _th_stunden_am_werktag(m, datum):
    """Echte Arbeitsstunden des TH am gegebenen Werktag aus arbeitszeit_gruppen[].Arbeitszeiten[].
    Wochentag-Codes als Bitmask: Mo=1, Di=2, Mi=4, Do=8, Fr=16.
    Gruppen-Gültigkeit über GueltigAb/GueltigBis. Mehrere Slots pro Tag werden summiert."""
    if datum.weekday() >= 5: return 0.0
    bitmask = 1 << datum.weekday()
    iso = datum.isoformat()
    total = 0.0
    for g in (m.get('arbeitszeit_gruppen') or []):
        # Gruppen-Gültigkeit
        g_von = g.get('GueltigAb'); g_bis = g.get('GueltigBis')
        if g_von and g_von > iso: continue
        if g_bis and g_bis < iso: continue
        for s in (g.get('Arbeitszeiten') or []):
            if s.get('Wochentag') != bitmask: continue
            if s.get('GueltigAb') and s['GueltigAb'] > iso: continue
            try:
                sh, sm, _ = s['Start'].split(':')
                eh, em, _ = s['Ende'].split(':')
                total += (int(eh) + int(em)/60) - (int(sh) + int(sm)/60)
            except Exception: continue
    return total

def compute_quartal(pm, q_start, q_end, today=None):
    """Berechnet IST/Vstd/Abw/Feiertage/verfueg/eur60 für ein beliebiges Quartalsfenster.

    Eine gemeinsame Funktion für:
    - Live (q_end = aktuelles Q-Ende, today < q_end): `effective_end = today`, rechnet Q-bisher.
    - Q-End (q_end = letzter Q-Tag, today ≥ q_end): `effective_end = q_end`, rechnet komplettes Q.
    - Probezeit-Override: wenn `pm['startdatum']` weniger als 6 Monate vor q_end → Stufe=1.

    Konsistenz-Anker: `eff_days` pro TH = max(q_start, TH-Start+29) … min(effective_end, TH-Bis).
    Alle Größen (Vstd, Abw, Feiertage, IST) werden über dasselbe eff_days-Fenster gerechnet —
    damit ist die 29-Tage-Sperre strukturell eingebaut (Zähler+Nenner schrumpfen proportional).

    Return-Dict: q_start, q_end, effective_end, wochen, bundle_th_count, bundle_h_pro_woche,
    vstd_ber, abw_ber, feiertage_ber, verfueg, ist, eur60, tats_stufe, termine_count,
    termine_skip_29d, probezeit_aktiv.

    Return None wenn vstd_ber ≤ 0 oder verfueg ≤ 0 (Quartal noch nicht ausreichend bewertbar).
    """
    from datetime import date as _date, timedelta as _td
    today = today or _date.today()
    effective_end = min(q_end, today)
    if effective_end < q_start:
        return None

    iso_q_start = q_start.isoformat()
    iso_eff_end = effective_end.isoformat()
    iso_29ago_end = (effective_end - _td(days=29)).isoformat()

    bundle_standorte = [s.strip().lower().replace(' ', '_') for s in pm['bundle_standorte'].split(',')]

    # Probezeit-Check (PM-Vertrag § 8 Abs. 3) — gemeinsame Helper-Funktion
    probezeit_aktiv = is_probezeit(pm.get('startdatum'), q_end)

    # Bundle-TH: Beschäftigung überlappt mit Q-Fenster UND Start ≤ effective_end - 29 Tage
    ma = _fetch_all('mc934lbrlg7w6e1')
    def _aktiv_im_q(m):
        bz = m.get('beschaeftigungszeiten') or []
        for e in bz:
            v = e.get('Von'); b = e.get('Bis')
            if (v is None or v <= iso_eff_end) and (b is None or b >= iso_q_start):
                return True
        return False
    def _seit_29tage_vor_eff_end(m):
        v = _th_earliest_beschaeftigung(m)
        return v is not None and v <= iso_29ago_end

    bundle_th = [m for m in ma
                 if m.get('is_therapeut')
                 and 'Online' not in f"{m.get('vorname','')} {m.get('nachname','')}"
                 and any(f in bundle_standorte for f in (m.get('filialen') or []))
                 and _aktiv_im_q(m)
                 and _seit_29tage_vor_eff_end(m)]
    th_ids = {m['id'] for m in bundle_th}
    th_by_id = {m['id']: m for m in bundle_th}
    th_start_iso = {m['id']: _th_earliest_beschaeftigung(m) for m in bundle_th}
    # Anlauf-TH (Start < 3 Monate): 4W-arbeitszeit_h ist nur ein Teilfenster-Wert und seit dem
    # n8n-Bridging-Update nicht mehr 0 -> für die Wochenstunden den StundenProWoche-Fallback nutzen.
    bridge_cutoff = _anlauf_cutoff(today)

    # Wochenstunden pro TH: primär auslastung_4w.arbeitszeit_h/4, Fallback StundenProWoche
    ausl = _fetch_all('m29vw64nhicfco2')
    latest_per_th = {}
    for r in ausl:
        mid = r.get('mitarbeiter_id')
        if mid not in th_ids: continue
        d = r.get('datum', '')
        if mid not in latest_per_th or d > latest_per_th[mid].get('datum', ''):
            latest_per_th[mid] = r

    # eff_days und Vstd pro TH
    bundle_h_pro_woche = 0.0
    vstd_ber = 0.0
    th_eff_start = {}   # für IST/Abw/Feiertage-Filter
    th_eff_end = {}
    for m in bundle_th:
        snap = latest_per_th.get(m['id'])
        h_woche = ((snap.get('arbeitszeit_h', 0) or 0) / 4) if (snap and not _ist_anlauf_th(m, bridge_cutoff)) else 0
        if h_woche <= 0:
            g = (m.get('arbeitszeit_gruppen') or [{}])[0]
            h_woche = float(g.get('StundenProWoche', 0) or 0)
        bundle_h_pro_woche += h_woche

        eff_start = q_start
        start_iso = th_start_iso.get(m['id'])
        if start_iso:
            try:
                sperre_ende = _date.fromisoformat(start_iso) + _td(days=29)
                eff_start = max(q_start, sperre_ende)
            except Exception: pass

        eff_end = effective_end
        for e in (m.get('beschaeftigungszeiten') or []):
            if e.get('Bis'):
                try:
                    bis_d = _date.fromisoformat(e['Bis'])
                    if bis_d < eff_end:
                        eff_end = bis_d
                except Exception: pass

        if eff_start > eff_end:
            th_eff_start[m['id']] = None
            th_eff_end[m['id']] = None
            continue

        th_eff_start[m['id']] = eff_start
        th_eff_end[m['id']] = eff_end
        eff_days = (eff_end - eff_start).days + 1
        vstd_ber += h_woche * eff_days / 7

    if vstd_ber <= 0:
        return None

    # IST aus Termine (Termin im TH-eff_days-Range)
    ist = 0.0
    termine_count = 0
    termine_skip_29d = 0
    for st in bundle_standorte:
        termine = _fetch_all('mf2pw17nwfzlkd2', where=f'(filiale,eq,{st})')
        for t in termine:
            if t.get('deleted_at'): continue
            if t.get('art') != 'normal': continue
            if t.get('is_blocker') or t.get('is_passive_leistung'): continue
            if t.get('status') not in ('erbracht', 'erbracht_und_unterschrieben'): continue
            try:
                b = _date.fromisoformat(t['beginn'][:10])
            except Exception: continue
            ma_list = t.get('mitarbeiter') or []
            if not ma_list: continue
            mid = ma_list[0].get('Id')
            if mid not in th_ids: continue
            es = th_eff_start.get(mid); ee = th_eff_end.get(mid)
            if es is None or ee is None: continue
            if b < es:
                termine_skip_29d += 1
                continue
            if b > ee: continue
            ist += termin_umsatz(t)
            termine_count += 1

    # VO-Gebühren: 10 € Blattgebühr je VO + 98,59 € je Blanko-VO, zugeordnet dem
    # Quartal des LETZTEN VO-Termins (über alle Status, wie Monatsumsatz-Report V3).
    # Kein TH-Fenster-Filter — Gebühren sind VO-Ebene, nicht Termin-Ebene.
    vo_last = {}
    for st in bundle_standorte:
        for t in _fetch_all('mf2pw17nwfzlkd2', where=f'(filiale,eq,{st})'):
            if t.get('deleted_at') or t.get('art') != 'normal' or t.get('is_blocker'):
                continue
            vo = str(t.get('verordnung_id') or '')
            if not vo: continue
            try:
                b = _date.fromisoformat(t['beginn'][:10])
            except Exception:
                continue
            cur = vo_last.get(vo)
            if not cur or b >= cur['datum']:
                vo_last[vo] = {'datum': b, 'blanko': (cur['blanko'] if cur else False) or bool(t.get('is_blanko'))}
            elif t.get('is_blanko'):
                cur['blanko'] = True
    vo_gebuehren = 0.0
    for vo in vo_last.values():
        if q_start <= vo['datum'] <= effective_end:
            vo_gebuehren += VO_BLATTGEBUEHR
            if vo['blanko']:
                vo_gebuehren += BLANKO_PAUSCHALE * satz_faktor(vo['datum'].isoformat())
    ist += vo_gebuehren

    # Abw_ber: individuell pro Wochentag, eff_days-Range pro TH
    EXCLUDED_ARTS = {'krank', 'krankheit_kind', 'angefragt'}
    abw_records = _fetch_all('mwcnx74etcl1frq')
    abw_ber = 0.0
    for a in abw_records:
        if a.get('deleted_at'): continue
        if a.get('art') in EXCLUDED_ARTS: continue
        mid = a.get('mitarbeiter_id')
        if mid not in th_ids: continue
        es = th_eff_start.get(mid); ee = th_eff_end.get(mid)
        if es is None or ee is None: continue
        m_th = th_by_id[mid]
        try:
            von = _date.fromisoformat(a['von'][:10])
            bis = _date.fromisoformat(a['bis'][:10])
        except Exception: continue
        day = max(von, es); end_day = min(bis, ee)
        while day <= end_day:
            if day.weekday() < 5:
                abw_ber += _th_stunden_am_werktag(m_th, day)
            day += _td(days=1)

    # Feiertage_ber: pro Werktag-Feiertag, eff_days-Range pro TH
    feiertage_ber = 0.0
    day = q_start
    while day <= effective_end:
        iso = day.isoformat()
        if iso in BERLIN_FEIERTAGE and day.weekday() < 5:
            for m in bundle_th:
                es = th_eff_start.get(m['id']); ee = th_eff_end.get(m['id'])
                if es is None or ee is None: continue
                if day < es or day > ee: continue
                feiertage_ber += _th_stunden_am_werktag(m, day)
        day += _td(days=1)

    verfueg = vstd_ber - abw_ber - feiertage_ber
    if verfueg <= 0:
        return None
    eur60 = ist / verfueg

    # rechn_stufe = rein aus eur60 + zufr (UND-Logik), OHNE Übersprungs-Limit
    rechn_stufe = 1
    for s in reversed(STUFEN):
        if eur60 >= s['eur60'] and pm.get('zufr', 0) >= s['zufr']:
            rechn_stufe = s['n']; break

    # tats_stufe = mit Übersprungs-Limit gegenüber start_stufe (Vorquartal-Stufe)
    # + Probezeit-Override (fix Stufe 1)
    start_stufe = pm.get('start_stufe', rechn_stufe)
    if probezeit_aktiv:
        tats_stufe = 1
    else:
        tats_stufe = max(start_stufe - MAX_STUFEN_SPRUNG,
                         min(rechn_stufe, start_stufe + MAX_STUFEN_SPRUNG))
        tats_stufe = max(1, min(tats_stufe, 6))

    wochen = ((effective_end - q_start).days + 1) / 7
    return {
        'q_start': q_start, 'q_end': q_end, 'effective_end': effective_end,
        'wochen': wochen,
        'bundle_th_count': len(bundle_th),
        'anzahl_th_aktiv': len(bundle_th),
        'bundle_h_pro_woche': bundle_h_pro_woche,
        'vstd_ber': vstd_ber,
        'abw_ber': abw_ber,
        'feiertage_ber': feiertage_ber,
        'verfueg': verfueg,
        'ist': ist,
        'vo_gebuehren': vo_gebuehren,
        'eur60': eur60,
        'rechn_stufe': rechn_stufe,   # rechnerisch erreichte Stufe (motivierender Wert)
        'tats_stufe': tats_stufe,     # mit ±1-Deckel + Probezeit-Override (bewertungsrelevant)
        'termine_count': termine_count,
        'termine_skip_29d': termine_skip_29d,
        'probezeit_aktiv': probezeit_aktiv,
    }


def compute_live_quartalsstand(pm, today=None):
    """Wrapper: delegiert an compute_quartal() für das laufende Quartal.

    Behält das alte Return-Schema (eur60_live, abw_h_gemessen, ...) für UI-Kompatibilität.
    """
    from datetime import date as _date, timedelta as _td
    today = today or _date.today()
    q_month = ((today.month - 1) // 3) * 3 + 1
    q_start = _date(today.year, q_month, 1)
    # Q-Ende ermitteln
    if q_month + 3 > 12:
        q_end = _date(today.year + 1, 1, 1) - _td(days=1)
    else:
        q_end = _date(today.year, q_month + 3, 1) - _td(days=1)

    # Live-Aufruf: start_stufe für ±1-Deckel = tats_stufe des Bewertungs-Quartals (= Q1)
    pm_live = dict(pm)
    pm_live['start_stufe'] = pm.get('tats_stufe', pm.get('start_stufe', 1))
    result = compute_quartal(pm_live, q_start, q_end, today)
    if not result:
        return None

    # Map auf altes Schema für Backwards-Kompatibilität
    q1_abw_quote = pm['abw_ber'] / pm['vstd_ber'] if pm.get('vstd_ber', 0) > 0 else 0.10
    return {
        'eur60_live': result['eur60'],
        'ist_live': result['ist'],
        'verfueg_live': result['verfueg'],
        'vstd_q_bisher': result['vstd_ber'],
        'abw_h_stabilisiert': result['vstd_ber'] * q1_abw_quote,
        'abw_h_gemessen': result['abw_ber'],
        'feiertage_h_gemessen': result['feiertage_ber'],
        'q1_abw_quote': q1_abw_quote,
        'bundle_h_pro_woche': result['bundle_h_pro_woche'],
        'wochen_q_bisher': result['wochen'],
        'anzahl_th_aktiv': result['anzahl_th_aktiv'],
        'termine_count': result['termine_count'],
        'termine_skip_29d': result['termine_skip_29d'],
        'q_start': result['q_start'],
        'today': result['effective_end'],
        'tats_stufe_live': result['tats_stufe'],         # mit ±1-Deckel
        'rechn_stufe_live': result.get('rechn_stufe', result['tats_stufe']),  # ohne Deckel
        'probezeit_aktiv': result.get('probezeit_aktiv', False),
    }



def compute_live_kpis(bundle_standorte, today=None):
    """Berechnet Auslastung, PKV-Quote, Krank-Tage/TH/Jahr für Bundle."""
    from datetime import date as _date
    today = today or _date.today()
    # Quartalsstart
    q_month = ((today.month - 1) // 3) * 3 + 1
    q_start = _date(today.year, q_month, 1)
    
    # 1) TH im Bundle
    ma = _fetch_all('mc934lbrlg7w6e1')
    bundle_th = [m for m in ma 
                 if m.get('is_therapeut')
                 and 'Online' not in f"{m.get('vorname','')} {m.get('nachname','')}"
                 and any(f in bundle_standorte for f in (m.get('filialen') or []))]
    th_ids = {t['id'] for t in bundle_th}
    
    # 2) AUSLASTUNG: aus Auslastung 4W Tabelle (rolling 30 Tage)
    auslast_records = _fetch_all(AUSLASTUNG_4W_TABLE)
    # Anlauf-Therapeuten (Start < 3 Monate, = 4W-Schwelle minMonateAktiv) ausschließen –
    # konsistent zur App (include_in_site=false). NocoDB hat das Flag nicht, daher über das
    # Beschäftigungs-Startdatum. Vorher trugen sie 0/0 bei (neutral); seit dem n8n-Bridging-
    # Update liefern sie echte Anlaufwerte – ohne diesen Filter würden sie die KPI verfälschen.
    bridge_cutoff = _anlauf_cutoff(today)
    neu_th_ids = {t['id'] for t in bundle_th if _ist_anlauf_th(t, bridge_cutoff)}
    # Nimm letzten Snapshot pro TH (Anlauf-TH übersprungen)
    latest_per_th = {}
    for r in auslast_records:
        mid = r.get('mitarbeiter_id')
        if mid not in th_ids: continue
        if mid in neu_th_ids: continue
        d = r.get('datum','')
        if mid not in latest_per_th or d > latest_per_th[mid].get('datum',''):
            latest_per_th[mid] = r
    # Bundle-Auslastung wie in der App (AdminAuslastung.tsx): SUM(ist_h) / SUM(ziel_h)
    # — gewichtet nach Wochenstunden, nicht avg der Prozent-Werte
    total_ist  = sum(r.get('ist_h', 0) or 0 for r in latest_per_th.values())
    total_ziel = sum(r.get('zielwert_h', 0) or 0 for r in latest_per_th.values())
    auslastung = (total_ist / total_ziel * 100) if total_ziel > 0 else None
    
    # 3) PKV-QUOTE: aus Termine Q-bisher
    pkv_count = 0; total_count = 0
    for st in bundle_standorte:
        termine = _fetch_all('mf2pw17nwfzlkd2', where=f'(filiale,eq,{st})')
        for t in termine:
            if t.get('deleted_at'): continue
            from datetime import date as _dt
            try: 
                b = _dt.fromisoformat(t['beginn'][:10])
            except: continue
            if b < q_start or b > today: continue
            if t.get('status') not in ('erbracht','erbracht_und_unterschrieben'): continue
            if t.get('art') != 'normal': continue
            total_count += 1
            if t.get('verordnungstyp') in (2, 3):
                pkv_count += 1
    pkv_quote = pkv_count / total_count * 100 if total_count else None
    
    # 4) KRANKHEIT: trailing 90 Kalendertage, hochgerechnet aufs Jahr
    # — rolling, damit der Wert beim Quartalswechsel nicht abrupt zurückspringt
    #   und Spitzenmonate (z.B. Februar-Welle) im Mess-Fenster bleiben
    from datetime import timedelta as _td
    from datetime import date as _dt
    krank_start = today - _td(days=90)
    abw = _fetch_all('mwcnx74etcl1frq')
    krank_stunden = 0.0
    for a in abw:
        if a.get('deleted_at'): continue
        if a.get('art') not in ('krank','krankheit_kind'): continue
        if a.get('mitarbeiter_id') not in th_ids: continue
        try:
            von = _dt.fromisoformat(a['von'][:10])
            bis = _dt.fromisoformat(a['bis'][:10])
        except: continue
        day = max(von, krank_start); end_day = min(bis, today)
        while day <= end_day:
            if day.weekday() < 5:
                # Hours für Standard-TH (grob 8h, vereinfacht)
                krank_stunden += 8.0
            day += _td(days=1)
    # Werktage im 90-Tage-Fenster
    werktage_fenster = 0
    d = krank_start
    while d <= today:
        if d.weekday() < 5: werktage_fenster += 1
        d += _td(days=1)
    krank_tage_bundle_jahr = (krank_stunden / 8) * (230 / werktage_fenster) if werktage_fenster else 0
    anzahl_th = len(bundle_th)
    krank_tage_pro_th_jahr = krank_tage_bundle_jahr / anzahl_th if anzahl_th else 0
    
    return {
        'auslastung': auslastung,
        'auslastung_n_th': len(latest_per_th),
        'pkv_quote': pkv_quote,
        'pkv_termine_total': total_count,
        'krank_tage_pro_th_jahr': krank_tage_pro_th_jahr,
        'q_start': q_start,
        'today': today,
        'werktage_fenster': werktage_fenster,
    }

def level_auslastung(val):
    if val is None: return None
    if val < 85: return 'low'
    if val < 93: return 'mid'
    if val < 96: return 'high'
    return 'vhigh'

def level_pkv(val):
    if val is None: return None
    if val <= 10: return 'low'
    if val <= 20: return 'mid'
    if val <= 30: return 'high'
    return 'vhigh'

def level_krank(val):
    """Krank-Tage/TH/Jahr: ≤15=low, 16–20=mid, 21–25=high, >25=vhigh"""
    if val is None: return None
    if val <= 15: return 'low'
    if val <= 20: return 'mid'
    if val <= 25: return 'high'
    return 'vhigh'

def kpi_level_label(kpi, level):
    if kpi == 'Auslastung':
        return {'low':'niedrig','mid':'mittel','high':'hoch','vhigh':'sehr hoch'}.get(level, '—')
    if kpi == 'PKV-Quote':
        return {'low':'gering','mid':'mittel','high':'hoch','vhigh':'sehr hoch'}.get(level, '—')
    if kpi == 'Krankheit':
        return {'low':'wenig','mid':'normal','high':'höher'}.get(level, '—')
    return '—'

def render_html(pm):
    from datetime import date as _date
    today = _date.today()
    q_aktuell = quartal_label(today)        # laufendes Quartal, z.B. 'Q2 2026'
    q_bewertung = vorquartal_label(today)   # zuletzt bewertetes Quartal, z.B. 'Q1 2026'
    q_now_num = (today.month - 1) // 3 + 1  # 1..4 — fürs Timeline-Mapping

    d = delta_naechste_stufe(pm)

    live_kpis = None  # gefüllt nur wenn nicht Stufe 6 (siehe unten)

    # Hero
    hero_stufe_text = f"Stufe {pm['tats_stufe']}"
    hero_probezeit_note = ''
    einordnung_probezeit_note = ''
    if pm.get('probezeit_aktiv'):
        hero_stufe_text += ' · Probezeit'
        hero_probezeit_note = ('<div class="hero-meta" style="margin-top:10px;font-size:13px;opacity:0.92;">'
                               'Probezeit-Regel: In den ersten 6 Monaten gilt fest Stufe 1 — '
                               'unabhängig von Umsatz und Zufriedenheit. Deine erste reguläre '
                               'Bewertung folgt im Quartal nach Probezeit-Ende.</div>')
        einordnung_probezeit_note = ('<p style="margin:0 0 14px;padding:10px 14px;background:rgba(13,89,90,0.08);'
                                     'border-radius:8px;font-size:13.5px;">Hinweis: Du bist noch in der '
                                     'Probezeit-Regel — deine Stufe ist deshalb fest auf 1 gesetzt, auch wenn '
                                     'deine Werte unten bereits höhere Schwellen erreichen. Ab der ersten '
                                     'regulären Bewertung zählen sie ganz normal.</p>')

    # Breakdown — Segment-Anteile
    ziel_gehalt = max(pm['gehalt_formel'], pm['mindest_anteilig'])
    total_width = pm['jahresgehalt']
    sockel_pct = (40000 * pm['wochenstd']/40) / total_width * 100
    bundle_pct = (pm['bundle_zulage'] * pm['wochenstd']/40) / total_width * 100
    stufe_pct = max(0, 100 - sockel_pct - bundle_pct)
    
    # Stufen-Leiter
    stufen_chips = []
    for s in STUFEN:
        cls = ''
        if s['n'] == pm['tats_stufe']: cls = 'current'
        elif s['n'] == pm['tats_stufe'] + 1: cls = 'next'
        stufen_chips.append(f'''
        <div class="stufe-chip {cls}">
          <div class="stufe-chip-num">{"★ Deine Stufe" if cls == "current" else ("Nächste Stufe" if cls == "next" else "Stufe")}</div>
          <div class="stufe-chip-name">{s['n']}</div>
          <div class="stufe-chip-detail">€/h ≥ {fmt_eur(s['eur60'], 2)}<br>Zufriedenheit ≥ {fmt_de(s['zufr'])}<br>+{int(s['zulage']*100)} % Zulage</div>
        </div>''')
    
    # €/60min-Block: Schwelle aktuell + progress
    curr_s = STUFEN[pm['tats_stufe']-1]
    next_s = STUFEN[min(pm['tats_stufe'], 5)]
    eur_progress = max(0, min(100, (pm['eur60'] - curr_s['eur60']) / 
                                   (next_s['eur60'] - curr_s['eur60']) * 100 if next_s != curr_s else 100))
    
    # Zufr
    zufr_curr = curr_s['zufr']
    zufr_next = next_s['zufr']
    zufr_progress = 100 if zufr_next == zufr_curr else max(0, min(100, (pm['zufr'] - zufr_curr) / 
                                                                        (zufr_next - zufr_curr) * 100))
    
    # Nächste Stufe - Was bringt / was fehlt
    next_block = ''
    gap_block = ''
    wege_block_html = ''
    hebel_note_condition = ''
    if d:
        next_s_obj = d['next_stufe']
        next_block = f'''
        <div class="block">
          <div class="block-label">Ausblick</div>
          <div class="block-title">Was bringt dir Stufe {next_s_obj['n']}?</div>
          <div class="kpi-grid">
            <div class="kpi-card">
              <div class="kpi-label">Monatsgehalt bei Stufe {next_s_obj['n']}</div>
              <div class="kpi-value good">{fmt_eur(d['next_monatsgehalt'])} <span class="kpi-unit">€</span></div>
              <div class="kpi-schwelle">Δ zum jetzigen Gehalt: <strong>+{fmt_eur(d['delta_gehalt_monat'])} € / Monat</strong></div>
            </div>
            <div class="kpi-card">
              <div class="kpi-label">Jahresgehalt bei Stufe {next_s_obj['n']}</div>
              <div class="kpi-value good">{fmt_eur(d['next_jahresgehalt'])} <span class="kpi-unit">€</span></div>
              <div class="kpi-schwelle">Δ pro Jahr: <strong>+{fmt_eur(d['delta_gehalt_jahr'])} €</strong></div>
            </div>
          </div>
        </div>
        '''
        # Gap: Umsatz + Zufriedenheit separat prüfen
        umsatz_reicht = pm['eur60'] >= next_s_obj['eur60']
        zufr_reicht = pm['zufr'] >= next_s_obj['zufr']
        delta_zufr = next_s_obj['zufr'] - pm['zufr']
        eur_progress = max(0, min(100, (pm['eur60'] - curr_s['eur60']) / 
                                  (next_s_obj['eur60'] - curr_s['eur60']) * 100)) if next_s_obj['eur60'] > curr_s['eur60'] else 100
        zufr_progress = max(0, min(100, (pm['zufr'] - curr_s['zufr']) / 
                                   (next_s_obj['zufr'] - curr_s['zufr']) * 100)) if next_s_obj['zufr'] > curr_s['zufr'] else 100
        
        # Umsatz-Bereich (entweder Gap oder ✓)
        if umsatz_reicht:
            umsatz_html = f'''
            <div class="gap-row done">
              <div class="gap-row-icon">✓</div>
              <div class="gap-row-content">
                <div class="gap-row-title">Umsatz pro Therapie-Stunde — erreicht</div>
                <div class="gap-row-detail">Du hast {fmt_eur(pm['eur60'], 2)} €/h, Schwelle für Stufe {next_s_obj['n']}: {fmt_eur(next_s_obj['eur60'], 2)} €/h</div>
              </div>
            </div>'''
        else:
            umsatz_html = f'''
            <div class="gap-row gap">
              <div class="gap-row-icon">↗</div>
              <div class="gap-row-content">
                <div class="gap-row-title">Mehr Umsatz pro Therapie-Stunde</div>
                <div class="gap-row-kpis">
                  <div class="gap-mini">
                    <div class="gap-mini-label">Fehlt</div>
                    <div class="gap-mini-val">+{fmt_eur(d['delta_eur60'], 2)} €/h</div>
                  </div>
                  <div class="gap-mini">
                    <div class="gap-mini-label">Als Bundle-Umsatz/Monat</div>
                    <div class="gap-mini-val">+{fmt_eur(d['delta_bundle_monat'])} €</div>
                  </div>
                </div>
                <div class="gap-bar"><div class="gap-bar-fill" style="width: {eur_progress:.0f}%"></div></div>
                <div class="gap-bar-caption">{fmt_eur(pm['eur60'], 2)} €/h → {fmt_eur(next_s_obj['eur60'], 2)} €/h</div>
              </div>
            </div>'''
        
        # Zufriedenheits-Bereich
        if zufr_reicht:
            zufr_html = f'''
            <div class="gap-row done">
              <div class="gap-row-icon">✓</div>
              <div class="gap-row-content">
                <div class="gap-row-title">Team-Zufriedenheit — erreicht</div>
                <div class="gap-row-detail">Du hast {fmt_de(pm['zufr'])} / 10, Schwelle für Stufe {next_s_obj['n']}: {fmt_de(next_s_obj['zufr'])}</div>
              </div>
            </div>'''
        else:
            zufr_html = f'''
            <div class="gap-row gap">
              <div class="gap-row-icon">↗</div>
              <div class="gap-row-content">
                <div class="gap-row-title">Höhere Team-Zufriedenheit</div>
                <div class="gap-row-kpis">
                  <div class="gap-mini">
                    <div class="gap-mini-label">Fehlt</div>
                    <div class="gap-mini-val">+{fmt_de(delta_zufr)} Punkte</div>
                  </div>
                  <div class="gap-mini">
                    <div class="gap-mini-label">Ziel</div>
                    <div class="gap-mini-val">{fmt_de(next_s_obj['zufr'])} / 10</div>
                  </div>
                </div>
                <div class="gap-bar"><div class="gap-bar-fill" style="width: {zufr_progress:.0f}%"></div></div>
                <div class="gap-bar-caption">{fmt_de(pm['zufr'])} → {fmt_de(next_s_obj['zufr'])}</div>
              </div>
            </div>'''
        
        # Kombiniere
        alles_erreicht = umsatz_reicht and zufr_reicht
        bundle_umsatz_hint = ''
        if not alles_erreicht:
            # Zeige wie viele Termine oder was auch immer
            pass
        
        gap_block = f'''
        <div class="block">
          <div class="block-label">Weg zur nächsten Stufe</div>
          <div class="block-title">Wie kommst du auf Stufe {next_s_obj['n']}?</div>
          <p class="block-intro">Für Stufe {next_s_obj['n']} müssen <strong>beide</strong> Bedingungen erfüllt sein — Umsatz <strong>und</strong> Team-Zufriedenheit.</p>
          <div class="gap-card">
            {umsatz_html}
            {zufr_html}
          </div>
        </div>
        '''

        # Live-KPIs "Dein aktueller Stand"
        aktueller_stand_html = ''
        live_quartal = None
        try:
            bundle_std_list = [s.strip().lower().replace(' ', '_') for s in pm.get('bundle_standorte','').split(',')]
            # Normalize back mapping
            bundle_std_list = ['prenzlauer_berg' if s == 'prenzlauer_berg' else s for s in bundle_std_list]
            live_kpis = compute_live_kpis(bundle_std_list)
        except Exception as _e:
            print(f'    (Live-KPIs-Fehler {pm["name"]}: {_e})')
        try:
            live_quartal = compute_live_quartalsstand(pm)
        except Exception as _e:
            print(f'    (Live-Quartalsstand-Fehler {pm["name"]}: {_e})')

        if live_kpis:
            from datetime import date as _date
            today_str = live_kpis['today'].strftime('%d.%m.%Y')
            # Aufbauen
            rows_html = ''
            # Auslastung
            a_val = live_kpis['auslastung']
            a_lvl = level_auslastung(a_val) if a_val is not None else None
            a_label = kpi_level_label('Auslastung', a_lvl) if a_lvl else '—'
            a_text = f'{a_val:.0f} %' if a_val is not None else 'keine Daten'
            a_chip_pct = {'low':18,'mid':55,'high':78,'vhigh':100}.get(a_lvl, 50) if a_lvl else 0
            rows_html += f'''
            <div class="live-kpi-row">
              <div class="live-kpi-label">Auslastung</div>
              <div class="live-kpi-chip-track">
                <div class="live-kpi-chip weg-chip-{a_lvl or 'low'}" style="width:{a_chip_pct}%">
                  <span class="weg-chip-text">{a_label}</span>
                  <span class="weg-chip-range">{a_text}</span>
                </div>
              </div>
              <div class="live-kpi-note">letzte 30 Tage (Auslastungs-Workflow)</div>
            </div>'''
            # PKV
            p_val = live_kpis['pkv_quote']
            p_lvl = level_pkv(p_val) if p_val is not None else None
            p_label = kpi_level_label('PKV-Quote', p_lvl) if p_lvl else '—'
            p_text = f'{p_val:.0f} %' if p_val is not None else 'keine Daten'
            p_chip_pct = {'low':32,'mid':55,'high':78,'vhigh':100}.get(p_lvl, 50) if p_lvl else 0
            rows_html += f'''
            <div class="live-kpi-row">
              <div class="live-kpi-label">PKV-Quote</div>
              <div class="live-kpi-chip-track">
                <div class="live-kpi-chip weg-chip-{p_lvl or 'low'}" style="width:{p_chip_pct}%">
                  <span class="weg-chip-text">{p_label}</span>
                  <span class="weg-chip-range">{p_text}</span>
                </div>
              </div>
              <div class="live-kpi-note">{q_aktuell} bisher · Stichprobe: {fmt_eur(live_kpis['pkv_termine_total'])} Termine</div>
            </div>'''
            # Krankheit
            k_val = live_kpis['krank_tage_pro_th_jahr']
            k_lvl = level_krank(k_val) if k_val is not None else None
            k_label = kpi_level_label('Krankheit', k_lvl) if k_lvl else '—'
            k_text = f'{k_val:.0f} Tg./Jahr' if k_val is not None else '—'
            k_chip_pct = {'low':32,'mid':55,'high':78}.get(k_lvl, 50) if k_lvl else 0
            rows_html += f'''
            <div class="live-kpi-row">
              <div class="live-kpi-label">Krank-Tage/TH/Jahr</div>
              <div class="live-kpi-chip-track">
                <div class="live-kpi-chip weg-chip-{k_lvl or 'low'}" style="width:{k_chip_pct}%">
                  <span class="weg-chip-text">{k_label}</span>
                  <span class="weg-chip-range">{k_text}</span>
                </div>
              </div>
              <div class="live-kpi-note">letzte 90 Tage, aufs Jahr hochgerechnet</div>
            </div>'''

            # €/h Live mit Stufen-Tendenz
            # Doppelt: tats_stufe (mit ±1-Deckel) für Bewertungs-Tendenz,
            # rechn_stufe (rein aus eur60+zufr) als Motivations-Signal
            if live_quartal:
                _eur_live = live_quartal['eur60_live']
                _stufe_live = live_quartal['tats_stufe_live']
                _stufe_rechn = live_quartal.get('rechn_stufe_live', _stufe_live)
                _hinweis = ''
                if _stufe_rechn > _stufe_live:
                    _hinweis = f' <span style="color:var(--muted);font-size:11px;">(rechnerisch Stufe {_stufe_rechn} — durch ±1-Limit auf {_stufe_live} gedeckelt)</span>'
                # Tendenz-Label vs. Q1-Bewertung
                if _stufe_live > pm['tats_stufe']:
                    _tendenz = f'<span style="color:var(--green);font-weight:700;">↑ auf Kurs Richtung Stufe {_stufe_live}</span>{_hinweis}'
                    _chip_class = 'high'; _chip_pct = 88
                elif _stufe_live < pm['tats_stufe']:
                    _tendenz = f'<span style="color:var(--orange);font-weight:700;">↓ aktuell unter Q1-Niveau (Tendenz Stufe {_stufe_live})</span>{_hinweis}'
                    _chip_class = 'mid'; _chip_pct = 50
                else:
                    _tendenz = f'<span style="color:var(--teal);font-weight:700;">→ Stufe {_stufe_live} bestätigt</span>{_hinweis}'
                    _chip_class = 'high'; _chip_pct = 75
                rows_html += f'''
            <div class="live-kpi-row">
              <div class="live-kpi-label">€/h (Trend)</div>
              <div class="live-kpi-chip-track">
                <div class="live-kpi-chip weg-chip-{_chip_class}" style="width:{_chip_pct}%">
                  <span class="weg-chip-text">{_tendenz}</span>
                  <span class="weg-chip-range">{fmt_eur(_eur_live, 2)} €/h</span>
                </div>
              </div>
              <div class="live-kpi-note">Q-bisher ({live_quartal['wochen_q_bisher']:.1f} Wo, {live_quartal['termine_count']} Termine) · finale Bewertung am Q-Ende</div>
            </div>'''

            aktueller_stand_html = f'''
            <div class="block">
              <div class="block-label">Live</div>
              <div class="block-title">Wo stehst du aktuell ({q_aktuell})?</div>
              <p class="block-intro" style="font-size:12px;margin-bottom:14px;">Orientierung für das laufende Quartal — die finale Bewertung erfolgt nach Q-Abschluss aus den endgültigen Daten.</p>
              <div class="live-kpi-card">
                <div class="live-kpi-header">
                  <div class="live-status">Live</div>
                  <div class="live-kpi-date">Stand: {today_str}</div>
                </div>
                {rows_html}
              </div>
            </div>
            '''

        # Konkrete Wege (Kombinationen von Auslastung, PKV, Krankheit)
        wege_content = render_wege_block(next_s_obj['n'])
        wege_legende = '''
        <div class="weg-legende">
          <div class="weg-legende-title">Was bedeuten die Stufen?</div>
          <div class="weg-legende-row">
            <span class="weg-legende-kpi">Auslastung</span>
            <span class="weg-legende-item mid">mittel <span class="rg">85–92 %</span></span>
            <span class="weg-legende-item high">hoch <span class="rg">93–95 %</span></span>
            <span class="weg-legende-item vhigh">sehr hoch <span class="rg">96 %+</span></span>
          </div>
          <div class="weg-legende-row">
            <span class="weg-legende-kpi">PKV-Quote</span>
            <span class="weg-legende-item low">gering <span class="rg">bis 10 %</span></span>
            <span class="weg-legende-item mid">mittel <span class="rg">11–20 %</span></span>
            <span class="weg-legende-item high">hoch <span class="rg">21–30 %</span></span>
            <span class="weg-legende-item vhigh">sehr hoch <span class="rg">über 30 %</span></span>
          </div>
          <div class="weg-legende-row">
            <span class="weg-legende-kpi">Krankheit</span>
            <span class="weg-legende-item low">wenig <span class="rg">10–15 Tg./a</span></span>
            <span class="weg-legende-item mid">normal <span class="rg">16–20 Tg./a</span></span>
            <span class="weg-legende-item high">höher <span class="rg">21–25 Tg./a</span></span>
          </div>
        </div>
        '''
        # Zufriedenheits-Voraussetzung für die Wege (UND-Bedingung neben den 3 Variablen)
        if pm['zufr'] >= next_s_obj['zufr']:
            zufr_voraussetzung = f'<span style="color:var(--green);font-weight:700;">✓ erreicht ({fmt_de(pm["zufr"])} / 10)</span>'
        else:
            _delta_z = next_s_obj['zufr'] - pm['zufr']
            zufr_voraussetzung = f'<span style="color:var(--orange);font-weight:700;">noch +{fmt_de(_delta_z)} Pkt fehlen ({fmt_de(pm["zufr"])} → {fmt_de(next_s_obj["zufr"])})</span>'

        wege_block_html = f'''
        <div class="block">
          <div class="block-label">Konkrete Wege</div>
          <div class="block-title">Optionen um Stufe {next_s_obj['n']} zu erreichen</div>
          <p class="block-intro">Jeder Weg zeigt eine <strong>realistische</strong> Kombination — schon eine davon reicht für Stufe {next_s_obj['n']}. PKV-Quote ist überall „gering" gesetzt, weil das eurem aktuellen Stand entspricht; eine höhere PKV-Quote macht jeden Weg leichter (siehe Hebel oben).</p>
          <p class="block-intro" style="margin-top:-8px;">Voraussetzung in allen Wegen: Team-Zufriedenheit ≥ {fmt_de(next_s_obj['zufr'])} — {zufr_voraussetzung}.</p>
          {wege_content}
          {wege_legende}
        </div>
        '''

    # Block: Aktion „Wie kommst du auf Stufe N?" — Q-Start + Live-Tendenz + Hebel
    hebel_block_html = ''
    h = hebel_optionen(pm, d, live_kpis)
    # Live-Hebel: hebel_optionen mit eur60_live & verfueg_live aufrufen
    h_live = None
    d_live = None
    if h and live_quartal:
        pm_live_dict = dict(pm)
        pm_live_dict['eur60'] = live_quartal['eur60_live']
        pm_live_dict['verfueg'] = live_quartal['verfueg_live']
        d_live = delta_naechste_stufe(pm_live_dict)
        if d_live and d_live.get('delta_eur60', 0) > 0:
            h_live = hebel_optionen(pm_live_dict, d_live, live_kpis)
    if h:
        TAG_LABEL = {'realistic':'realistisch', 'borderline':'ambitioniert', 'impossible':'alleine nicht möglich'}
        def _tag(lvl): return f'<span class="hebel-tag {lvl}">{TAG_LABEL[lvl]}</span>'
        def _eff(lvl): return '' if lvl == 'realistic' else f' {lvl}'

        # Wenn Live-Hebel verfügbar: Werte daraus, plus Q-Start als Vergleich.
        # Sonst: Q1-Werte wie bisher.
        hsrc = h_live if h_live else h

        pkv_from_to = f'von {hsrc["pkv_now"]:.0f} % auf {hsrc["pkv_neu"]:.0f} %' if hsrc["pkv_now"] else f'auf ca. {hsrc["pkv_neu"]:.0f} %'
        if h_live and h["d_pkv_pkt"] != hsrc["d_pkv_pkt"]:
            pkv_from_to += f' <span style="color:var(--muted);font-size:11px;">(Q-Start: +{h["d_pkv_pkt"]:.0f} %-Pkt)</span>'

        if hsrc["d_termin_pro_th"] is not None and hsrc["anzahl_th"]:
            termin_from_to = f'Bundle-weit · ≈ +{fmt_de(hsrc["d_termin_pro_th"])} Termine/Wo pro Therapeut:in (im Bundle: {hsrc["anzahl_th"]} Therapeut:innen)'
        else:
            termin_from_to = 'Bundle-weit, zusätzlich zu heute'
        if h_live and h["d_termin_wo"] != hsrc["d_termin_wo"]:
            termin_from_to += f' <span style="color:var(--muted);font-size:11px;">(Q-Start: +{h["d_termin_wo"]:.0f}/Wo)</span>'

        krank_unmöglich = bool(hsrc["krank_now"] and hsrc["d_krank_tage"] > hsrc["krank_now"])
        if krank_unmöglich:
            krank_from_to = f'Krankenstand bereits niedrig ({hsrc["krank_now"]:.0f} Tg./TH/Jahr) — als alleiniger Hebel nicht ausreichend.'
        elif hsrc["krank_now"]:
            krank_from_to = f'von {hsrc["krank_now"]:.0f} auf {hsrc["krank_neu"]:.0f} Tg./TH/Jahr'
        else:
            krank_from_to = f'auf ca. {hsrc["krank_neu"]:.0f} Tg./TH/Jahr'
        krank_effect_text = 'reicht alleine nicht' if krank_unmöglich else f'−{hsrc["d_krank_tage"]:.0f} Tage/TH/Jahr'

        # Umsatz-Card: Q-Start vs. Live (zwei Spalten)
        if live_quartal:
            _next_eur = next_s_obj['eur60']
            _live_eur = live_quartal['eur60_live']
            _live_stufe = live_quartal['tats_stufe_live']
            _live_delta_zu_ziel = max(0, _next_eur - _live_eur)
            if _live_eur >= _next_eur:
                _live_color = 'var(--green)'
                _live_text = f'✓ über Schwelle Stufe {next_s_obj["n"]} — Live-Tendenz {_live_stufe}'
            elif _live_eur > pm['eur60']:
                _live_color = 'var(--teal)'
                _live_text = f'fehlen noch +{fmt_eur(_live_delta_zu_ziel, 2)} €/h zu Stufe {next_s_obj["n"]}'
            else:
                _live_color = 'var(--orange)'
                _live_text = f'aktuell unter Q1 — Trend Stufe {_live_stufe}'
            umsatz_card = f'''
      <div class="gap-row" style="border-top:none;padding-top:4px;">
        <div class="gap-row-icon" style="background:var(--teal-light);color:var(--teal);">€</div>
        <div class="gap-row-content">
          <div class="gap-row-title">Umsatz pro Therapie-Stunde</div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:8px;">
            <div>
              <div class="gap-mini-label">Bei Q-Start (Q1-Bewertung)</div>
              <div class="gap-mini-val" style="color:var(--teal);">{fmt_eur(pm['eur60'], 2)} €/h</div>
              <div style="font-size:11px;color:var(--muted);margin-top:4px;">fehlten +{fmt_eur(d['delta_eur60'], 2)} €/h zu Stufe {next_s_obj['n']} ({fmt_eur(_next_eur, 2)} €/h)</div>
            </div>
            <div>
              <div class="gap-mini-label">Live (Q-bisher · {live_quartal['wochen_q_bisher']:.1f} Wo)</div>
              <div class="gap-mini-val" style="color:{_live_color};">{fmt_eur(_live_eur, 2)} €/h</div>
              <div style="font-size:11px;color:{_live_color};margin-top:4px;font-weight:600;">{_live_text}</div>
            </div>
          </div>
        </div>
      </div>'''
        else:
            umsatz_card = umsatz_html

        # Zufriedenheit-Card mit Hinweis
        zufr_card = zufr_html.replace(
            '</div>\n            </div>',
            '</div>\n              <div style="font-size:11px;color:var(--muted);font-style:italic;margin-top:6px;">Wert aus Q1-Umfrage · Live-Update zur Zufriedenheit nicht möglich</div>\n            </div>',
            1
        )

        # Hebel-Headline mit Live-Vergleich
        if h_live:
            hebel_headline = f'Drei Hebel zur Umsatz-Lücke — bei Q-Start: <strong>+{fmt_de(h["delta_pct"])} %</strong>, aktuell live: <strong>+{fmt_de(h_live["delta_pct"])} %</strong> Bundle-Umsatz. Jeder Hebel allein würde reichen, eine Kombination ist meist realistischer (siehe Wege unten).'
        elif live_quartal and live_quartal['eur60_live'] >= next_s_obj['eur60']:
            hebel_headline = f'Live-Stand bereits über Schwelle Stufe {next_s_obj["n"]} — wenn das so weitergeht, wäre die nächste Stufe erreicht. Die Hebel unten zeigen den Q-Start-Stand zur Orientierung.'
        else:
            hebel_headline = f'Drei Hebel zur Umsatz-Lücke (<strong>+{fmt_de(h["delta_pct"])} % Bundle-Umsatz</strong>) — jeder einzeln würde reichen, eine Kombination ist meist realistischer (siehe Wege unten).'

        hebel_block_html = f'''
  <!-- BLOCK „Wie kommst du auf Stufe N?" — Vergleich Q-Start vs. Live -->
  <div class="block">
    <div class="block-label">Aktion</div>
    <div class="block-title">Wie kommst du auf Stufe {next_s_obj['n']}?</div>
    <p class="block-intro">Für Stufe {next_s_obj['n']} müssen <strong>beide</strong> Bedingungen erfüllt sein — Umsatz <strong>und</strong> Team-Zufriedenheit. <em style="color:var(--muted);">Live-Werte sind Orientierung; die finale Bewertung erfolgt am Q-Ende.</em></p>
    <div class="gap-card">
      {umsatz_card}
      {zufr_card}
    </div>
    <p class="hebel-headline" style="margin-top:24px">
      {hebel_headline}
    </p>
    <div class="hebel-grid">
      <div class="hebel-item">
        <div class="hebel-content">
          <div class="hebel-name">Höherer PKV-Anteil {_tag(hsrc["pkv_lvl"])}</div>
          <div class="hebel-desc">Privatpatienten aktiv ansprechen. PKV zahlt {fmt_de(PKV_FAKTOR)}× den GKV-Tarif.</div>
          <div class="hebel-from-to">{pkv_from_to}</div>
        </div>
        <div class="hebel-effect{_eff(hsrc["pkv_lvl"])}">+{hsrc["d_pkv_pkt"]:.0f} %-Pkt PKV</div>
      </div>
      <div class="hebel-item">
        <div class="hebel-content">
          <div class="hebel-name">Mehr Termine pro Woche {_tag(hsrc["termin_lvl"])}</div>
          <div class="hebel-desc">Auslastung erhöhen, neue Patienten gewinnen, Slots besser nutzen.</div>
          <div class="hebel-from-to">{termin_from_to}</div>
        </div>
        <div class="hebel-effect{_eff(hsrc["termin_lvl"])}">+{hsrc["d_termin_wo"]:.0f} Termine/Wo</div>
      </div>
      <div class="hebel-item">
        <div class="hebel-content">
          <div class="hebel-name">Krankenstand senken {_tag(hsrc["krank_lvl"])}</div>
          <div class="hebel-desc">Team-Gesundheit, gute Urlaubsplanung, keine Ansteckungsketten.</div>
          <div class="hebel-from-to">{krank_from_to}</div>
        </div>
        <div class="hebel-effect{_eff(hsrc["krank_lvl"])}">{krank_effect_text}</div>
      </div>
    </div>
    <p class="hebel-note">Lineare Näherungen — Live-Werte zeigen den Stand der ersten Q-Wochen und können sich mit kommenden Urlaubsblöcken noch verschieben. Faktoren: +1 %-Pkt PKV ≈ +{fmt_de(PKV_FAKTOR-1)} % Umsatz · +1 Termin/Wo Bundle ≈ +0,3 % · −1 %-Pkt Krank ≈ +0,5 %. PKV-Tarif {fmt_de(PKV_FAKTOR)}× GKV, 230 Werktage/Jahr.</p>
  </div>'''

    # Timeline: nur bewertete Quartale + laufendes (keine leeren Zukunfts-Karten)
    year = today.year
    q_eval = q_now_num - 1 if q_now_num > 1 else None  # None: Bewertung war Vorjahr-Q4 (selten)
    timeline_cards = []
    end_q = q_now_num  # bis und mit dem laufenden Quartal
    for q in range(1, end_q + 1):
        q_label_card = f"Q{q} {year}"
        if q == q_eval:
            timeline_cards.append(f'''<div class="timeline-item current">
        <div class="timeline-q">{q_label_card}</div>
        <div class="timeline-stufe">Stufe {pm['tats_stufe']}</div>
        <div class="timeline-gehalt">{fmt_eur(pm['monatsgehalt'])} € / Monat</div>
      </div>''')
        elif q == q_now_num:
            timeline_cards.append(f'''<div class="timeline-item empty">
        <div class="timeline-q">{q_label_card}</div>
        <div class="timeline-empty">Laufendes Quartal</div>
      </div>''')
        else:
            timeline_cards.append(f'''<div class="timeline-item empty">
        <div class="timeline-q">{q_label_card}</div>
        <div class="timeline-empty">bewertet</div>
      </div>''')
    timeline_html = '\n      '.join(timeline_cards)

    html_str = f'''<!DOCTYPE html>
<html lang="de">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Vacura — Dein Gehalt · {pm['name']}</title>
  <style>{CSS}</style>
</head>
<body>
<div class="container">
  
  <div class="topbar">
    <span class="topbar-name">Vacura · Gehalts-Dashboard</span>
    <span class="topbar-quartal">{q_bewertung}</span>
  </div>

  <h1 class="page-title">Hallo {pm['name']}.</h1>
  <p class="page-subtitle">Hier siehst du, wo du stehst und wie du auf die nächste Stufe kommst.</p>
  
  <!-- BLOCK 1: HERO -->
  <div class="block">
    <div class="hero-card">
      <div class="hero-stufe">{hero_stufe_text}</div>
      <div class="hero-gehalt">{fmt_eur(pm['monatsgehalt'])} €</div>
      <div class="hero-gehalt-unit">Monatsgehalt ab {q_aktuell}</div>
      {hero_probezeit_note}
      <div class="hero-meta">
        <div class="hero-meta-item">
          <span class="hero-meta-label">Jahresgehalt</span>
          <span class="hero-meta-value">{fmt_eur(pm['jahresgehalt'])} €</span>
        </div>
        <div class="hero-meta-item">
          <span class="hero-meta-label">Bundle</span>
          <span class="hero-meta-value">{pm['bundle_standorte']}</span>
        </div>
        <div class="hero-meta-item">
          <span class="hero-meta-label">Wochenstunden</span>
          <span class="hero-meta-value">{pm['wochenstd']} h</span>
        </div>
      </div>
    </div>
  </div>
  
  <!-- BLOCK 2: BREAKDOWN -->
  <div class="block">
    <div class="block-label">Transparenz</div>
    <div class="block-title">Wie setzt sich dein Gehalt zusammen?</div>
    <div class="breakdown-card">
      <div class="breakdown-formula">
        <strong>Sockel</strong> + <strong>Bundle-Zulage</strong> = <strong>Basis-Gehalt</strong>, davon <strong>+{int(pm['tats_stufe_zulage_pct']*100)} %</strong> Stufen-Zulage
      </div>
      <div class="breakdown-bar">
        <div class="breakdown-seg sockel" style="flex: {sockel_pct:.1f}">Sockel</div>
        <div class="breakdown-seg bundle" style="flex: {bundle_pct:.1f}">Bundle</div>
        <div class="breakdown-seg stufe" style="flex: {stufe_pct:.1f}">+ Stufe {pm['tats_stufe']}</div>
      </div>
      <div class="breakdown-legend">
        <div class="breakdown-legend-item">
          <div class="breakdown-legend-label">Sockel (Vollzeit)</div>
          <div class="breakdown-legend-val">{fmt_eur(40000)} €</div>
        </div>
        <div class="breakdown-legend-item bundle">
          <div class="breakdown-legend-label">Bundle-Zulage (für {pm['th_pm']} Anteile<sup>¹</sup>)</div>
          <div class="breakdown-legend-val">+{fmt_eur(pm['bundle_zulage'])} €</div>
        </div>
        <div class="breakdown-legend-item stufe">
          <div class="breakdown-legend-label">Variable Zulage (+{int(pm['tats_stufe_zulage_pct']*100)} %)</div>
          <div class="breakdown-legend-val">× {fmt_de(1 + pm['tats_stufe_zulage_pct'], 2)}</div>
        </div>
      </div>
      <div class="breakdown-result">
        <span class="breakdown-result-label">→ Jahresgehalt (bei {pm['wochenstd']} h/Woche)</span>
        <span class="breakdown-result-val">{fmt_eur(pm['jahresgehalt'])} €</span>
      </div>
      <p style="font-size:11px;color:var(--muted);margin-top:14px;line-height:1.5;">
        <sup>¹</sup> „Anteile" = anteilige Bundle-Größe für deine Zulage — gewichtet auf deine Wochenstunden und deinen Anteil am Bundle-Team.
      </p>
    </div>
  </div>

  <!-- BLOCK 3: WARUM DIESE STUFE -->
  <div class="block">
    <div class="block-label">Einordnung</div>
    <div class="block-title">Woher kommt deine Stufe?</div>
    {einordnung_probezeit_note}
    <p class="block-intro">Beide Werte müssen gleichzeitig die Schwelle erreichen. Umsatz und Team-Stimmung sind UND-verknüpft.</p>
    <div class="kpi-grid">
      <div class="kpi-card">
        <div class="kpi-label">Dein Umsatz pro Therapie-Stunde</div>
        <div class="kpi-value ok">{fmt_eur(pm['eur60'], 2)} <span class="kpi-unit">€/h</span></div>
        <div class="kpi-schwelle">Schwelle Stufe {pm['tats_stufe']}: <strong>{fmt_eur(curr_s['eur60'], 2)} €/h</strong>{' · Stufe ' + str(next_s['n']) + ': ' + fmt_eur(next_s['eur60'], 2) + ' €/h' if d else ''}</div>
        {_eur_bar_html(pm, curr_s, next_s, d)}
      </div>
      <div class="kpi-card">
        <div class="kpi-label">Team-Zufriedenheit</div>
        <div class="kpi-value ok">{fmt_de(pm['zufr'])} <span class="kpi-unit">/ 10</span></div>
        <div class="kpi-schwelle">Schwelle Stufe {pm['tats_stufe']}: <strong>{fmt_de(curr_s['zufr'])}</strong>{' · Stufe ' + str(next_s['n']) + ': ' + fmt_de(next_s['zufr']) if d else ''}</div>
        {_zufr_bar_html(pm, curr_s, next_s, d)}
        <div class="zufr-breakdown">
          <div class="zufr-breakdown-title">Zusammensetzung (gewichtete Summe der drei Dimensionen)</div>
          <div class="zufr-sub">
            <span class="zufr-sub-label">Rücken freihalten</span>
            <span class="zufr-sub-bar"><span class="zufr-sub-fill" style="width:{pm['ruecken']*10:.0f}%"></span></span>
            <span class="zufr-sub-val">{fmt_de(pm['ruecken'])} <small>· Gewicht 20 %</small></span>
          </div>
          <div class="zufr-sub">
            <span class="zufr-sub-label">Kommunikation</span>
            <span class="zufr-sub-bar"><span class="zufr-sub-fill" style="width:{pm['komm']*10:.0f}%"></span></span>
            <span class="zufr-sub-val">{fmt_de(pm['komm'])} <small>· Gewicht 20 %</small></span>
          </div>
          <div class="zufr-sub">
            <span class="zufr-sub-label">Weiterempfehlung (eNPS)</span>
            <span class="zufr-sub-bar"><span class="zufr-sub-fill" style="width:{pm['enps']*10:.0f}%"></span></span>
            <span class="zufr-sub-val">{fmt_de(pm['enps'])} <small>· Gewicht 60 %</small></span>
          </div>
          <div class="zufr-formel" style="font-size:11px;color:var(--muted);margin-top:10px;padding-top:8px;border-top:1px dashed var(--line);">
            Score = 0,2 × {fmt_de(pm['ruecken'])} + 0,2 × {fmt_de(pm['komm'])} + 0,6 × {fmt_de(pm['enps'])} = <strong>{fmt_de(pm['zufr'])}</strong>
          </div>
        </div>
      </div>
    </div>
  </div>
  
  <!-- BLOCK 4: STUFEN-LEITER -->
  <div class="block">
    <div class="block-label">Überblick</div>
    <div class="block-title">Die 6 Stufen</div>
    <div class="stufen-scroll">
      {"".join(stufen_chips)}
    </div>
  </div>
  
  {aktueller_stand_html}

  {next_block}

  {hebel_block_html}

  {wege_block_html}

  <!-- BLOCK 8: TIMELINE -->
  <div class="block">
    <div class="block-label">Verlauf</div>
    <div class="block-title">Deine Entwicklung</div>
    <div class="timeline-grid">
      {timeline_html}
    </div>
  </div>

  <div class="footer">
    Vacura · PM-Gehaltsmodell · Stand {q_bewertung}<br>
    <a href="#top">↑ Nach oben</a>
  </div>
  
</div>
</body>
</html>'''
    return html_str

# =====================================================================
# Q-END-ROUTINE
# Wird am 1. des Folge-Monats nach Q-Ende aufgerufen (1.7./1.10./1.1./1.4.)
# Befüllt Audit-Sheet, updated "Stufe Vorquartal" in Daten-Tab.
# =====================================================================

def parse_q_label(label):
    """'2026-Q2' → (date(2026,4,1), date(2026,6,30), 'Audit Q2 2026')."""
    from datetime import date as _date, timedelta as _td
    year, q = label.split('-Q')
    year = int(year); q = int(q)
    q_month_start = (q - 1) * 3 + 1
    q_start = _date(year, q_month_start, 1)
    if q == 4:
        q_end = _date(year, 12, 31)
    else:
        q_end = _date(year, q_month_start + 3, 1) - _td(days=1)
    audit_sheet_name = f'Audit Q{q} {year}'
    return q_start, q_end, audit_sheet_name


def previous_q_label(today=None):
    """Q-Label des letzten abgeschlossenen Quartals (für Q-End-Routine).

    am 1.7.2026 → '2026-Q2'   (Q2 wurde am 30.6. abgeschlossen)
    am 1.10.2026 → '2026-Q3'
    am 1.1.2027 → '2026-Q4'
    am 1.4.2027 → '2027-Q1'
    """
    from datetime import date as _date
    today = today or _date.today()
    q_now = (today.month - 1) // 3 + 1
    if q_now == 1:
        return f'{today.year - 1}-Q4'
    return f'{today.year}-Q{q_now - 1}'


# Historisch eingefrorene Quartale — werden niemals von Q-End-Routine überschrieben
Q_LABELS_EINGEFROREN = {'2026-Q1', '2026-Q2'}

def run_q_end_routine(wb, q_label):
    """Q-End-Berechnung: schreibt Werte ins Quartals-Bewertungen-Sheet (Long-Format).

    Stufe-Vorquartal wird automatisch aus der letzten tats_stufe der vorherigen Zeile
    in Quartals-Bewertungen abgeleitet (kein manuelles Update mehr nötig).

    Schutz vor Überschreiben eingefrorener Quartale: Q_LABELS_EINGEFROREN.
    """
    if q_label in Q_LABELS_EINGEFROREN:
        raise RuntimeError(
            f'{q_label} ist als historisch eingefroren markiert. '
            f'Quartals-Bewertungen-Zeile darf nicht überschrieben werden. '
            f'Wenn das wirklich gewollt ist: Q_LABELS_EINGEFROREN in generate.py editieren.'
        )
    q_start, q_end, _ = parse_q_label(q_label)
    # Sheet-Zeilen tragen das Anzeige-Format 'Q2 2026' (wie _find_qb_row/compute_pm),
    # der Routine-Parameter kommt aber als '2026-Q2' — ohne Normalisierung entstehen
    # Duplikat-Zeilen (passiert beim Q2-2026-Lauf am 08.07.).
    q_disp = _q_label_from_date(q_start)
    print(f'\n=== Q-End-Routine: {q_label} ({q_start} bis {q_end}) ===')

    if 'Quartals-Bewertungen' not in wb.sheetnames:
        raise RuntimeError('Sheet "Quartals-Bewertungen" fehlt im Excel — bitte Phase A laufen lassen')
    ws_qb = wb['Quartals-Bewertungen']

    results = []
    for pm_cfg in PMS:
        # Zufr-Score aus den Input-Spalten der Q-Zeile + Vorquartals-Stufe für den
        # ±1-Deckel an compute_quartal durchreichen — ohne das fällt zufr auf 0 zurück
        # und jede PM landet fälschlich auf rechn/tats Stufe 1 (Bug bis 22.07.2026).
        pm_cfg = dict(pm_cfg)
        qrow = _find_qb_row(ws_qb, q_disp, pm_cfg['name'])
        if qrow:
            vals = [ws_qb.cell(row=qrow, column=c).value for c in (3, 4, 5)]
            if all(isinstance(v, (int, float)) for v in vals):
                pm_cfg['zufr'] = vals[0] * 0.2 + vals[1] * 0.2 + vals[2] * 0.6
        prev_row = _find_qb_row(ws_qb, _previous_q_label_from(q_disp), pm_cfg['name'])
        if prev_row:
            prev_tats = ws_qb.cell(row=prev_row, column=15).value
            if isinstance(prev_tats, (int, float)):
                pm_cfg['start_stufe'] = int(prev_tats)
        result = compute_quartal(pm_cfg, q_start, q_end, today=q_end)
        if not result:
            print(f'  {pm_cfg["name"]}: compute_quartal lieferte None')
            continue
        results.append((pm_cfg, result))
        print(f'  {pm_cfg["name"]}: €/h={result["eur60"]:.2f}, rechn={result["rechn_stufe"]}, tats={result["tats_stufe"]}'
              + (' (Probezeit)' if result.get('probezeit_aktiv') else ''))

    # Werte in Quartals-Bewertungen schreiben
    from datetime import datetime as _dt
    now_str = _dt.now().strftime('%Y-%m-%d %H:%M')
    print(f'\nSchreibe Werte in Quartals-Bewertungen (Long-Format):')
    for pm_cfg, result in results:
        row = _find_qb_row(ws_qb, q_disp, pm_cfg['name'])
        if row is None:
            # Neue Zeile anhängen
            row = 8
            while ws_qb.cell(row=row, column=1).value:
                row += 1
            ws_qb.cell(row=row, column=1, value=q_disp)
            ws_qb.cell(row=row, column=2, value=pm_cfg['name'])
        # Zufr-Score aus Q-bisher Werten (Input-Spalten 3-5)
        ruecken = ws_qb.cell(row=row, column=3).value
        komm    = ws_qb.cell(row=row, column=4).value
        enps    = ws_qb.cell(row=row, column=5).value
        zufr = None
        if all(isinstance(v, (int, float)) for v in (ruecken, komm, enps)):
            zufr = ruecken*0.2 + komm*0.2 + enps*0.6
            ws_qb.cell(row=row, column=7, value=round(zufr, 2))
        # Code-Output-Spalten 8-19
        ws_qb.cell(row=row, column=8, value=round(result['vstd_ber'], 1))
        ws_qb.cell(row=row, column=9, value=round(result['abw_ber'], 1))
        ws_qb.cell(row=row, column=10, value=round(result['feiertage_ber'], 1))
        ws_qb.cell(row=row, column=11, value=round(result['ist']))
        ws_qb.cell(row=row, column=12, value=round(result['verfueg'], 1))
        ws_qb.cell(row=row, column=13, value=round(result['eur60'], 2))
        ws_qb.cell(row=row, column=14, value=result['rechn_stufe'])
        ws_qb.cell(row=row, column=15, value=result['tats_stufe'])
        ws_qb.cell(row=row, column=16, value='Ja' if result.get('probezeit_aktiv') else 'Nein')
        # Diff zu MediFox
        mf = ws_qb.cell(row=row, column=6).value
        if isinstance(mf, (int, float)):
            diff = result['ist'] - mf
            ws_qb.cell(row=row, column=17, value=round(diff))
            diff_pct = (diff / mf) * 100 if mf else 0
            status = '✓ OK' if abs(diff_pct) < 2 else f'⚠️ Diff {diff_pct:+.1f} %'
        elif zufr is None:
            status = '⏳ Zufriedenheit fehlt'
        else:
            status = '⏳ MediFox-IST fehlt'
        ws_qb.cell(row=row, column=18, value=status)
        ws_qb.cell(row=row, column=19, value=now_str)
        print(f'  Row {row}: {pm_cfg["name"]} → eur60={result["eur60"]:.2f}, tats={result["tats_stufe"]}, status="{status}"')

    return results


# ==== MAIN ====
def _main():
    import argparse
    _ap = argparse.ArgumentParser()
    _ap.add_argument('--q-end', metavar='YYYY-QN', nargs='?', const='AUTO',
                     help='Q-End-Routine ausführen. Ohne Argument: vorheriges abgeschlossenes Q (für Cron-Auto-Trigger).')
    _ap.add_argument('--save-excel', action='store_true',
                     help='Excel mit Q-End-Werten überschreiben (sonst nur Dry-Run).')
    _ap.add_argument('--check-tarife', action='store_true',
                     help='Nur Tarif-Änderungen prüfen (letzte 7 Tage NocoDB) → JSON auf STDOUT, dann beenden.')
    _args, _ = _ap.parse_known_args()

    if _args.check_tarife:
        import json
        from datetime import date as _date
        # Stufen brauchen wir für die Schwellen-Empfehlung
        wb_tmp = openpyxl.load_workbook(EXCEL, data_only=False)
        load_stufen_aus_excel(wb_tmp)
        diff = _check_tarif_aenderungen(_date.today())
        if diff is None:
            print(json.dumps({'status': 'no_change'}, ensure_ascii=False))
        else:
            print(json.dumps({'status': 'change_detected', **diff}, indent=2, ensure_ascii=False))
        return
    if _args.q_end == 'AUTO':
        _args.q_end = previous_q_label()
        print(f'Q-End-Routine Auto-Modus: letztes abgeschlossenes Quartal = {_args.q_end}')

    print('Lade Excel...')
    wb = openpyxl.load_workbook(EXCEL, data_only=False)
    load_stufen_aus_excel(wb)
    pms_changed = load_pms_from_excel(wb)
    if pms_changed:
        wb.save(EXCEL)
        print('  💾 Excel mit auto-erzeugten Tokens/Defaults gespeichert.')
    print(f'  PMs aus Excel: {", ".join(p["name"] for p in PMS)} ({len(PMS)} aktive)')
    print(f'  Stufen-Schwellen: '
          + ', '.join(f'{s["n"]}={s["eur60"]:.2f}€/h@zufr{s["zufr"]:.1f}' for s in STUFEN))

    # Q-End-Routine (optional)
    if _args.q_end:
        run_q_end_routine(wb, _args.q_end)
        if _args.save_excel:
            wb.save(EXCEL)
            print(f'\n✅ Excel gespeichert: {EXCEL}')
        else:
            print('\n(Dry-Run — Excel nicht gespeichert. Mit --save-excel überschreiben.)')

    os.makedirs(OUT_DIR, exist_ok=True)

    from datetime import date as _date_main
    q_label_dash = vorquartal_label(_date_main.today())
    print(f'  Bewertungs-Quartal für Dashboards: {q_label_dash}')
    for pm_cfg in PMS:
        print(f'  {pm_cfg["name"]}...')
        pm_data = compute_pm(wb, pm_cfg, q_label=q_label_dash)
        if not pm_data:
            print(f'    übersprungen (keine Daten)')
            continue

        html_out = render_html(pm_data)
        token = TOKENS.get(pm_cfg["name"], "")
        out_path = os.path.join(OUT_DIR, f'{pm_cfg["name"].lower()}-{token}.html')
        with open(out_path, 'w') as f:
            f.write(html_out)
        print(f'    → {out_path}  (Stufe {pm_data["tats_stufe"]}, {fmt_eur(pm_data["monatsgehalt"])} €/Monat)')

    print('\nFertig.')


if __name__ == '__main__':
    _main()
